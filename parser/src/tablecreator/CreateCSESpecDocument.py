"""

The purpose of this document is to list out the specialization  
that are available for CSE students. A specialization for a 
CSE student is optional, unlike ISE students, where they are
required to take a specialization. 

These specializations will be on a separate sheet, apart from
the necessary information that is displayed on the first sheet 

"""


from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

# Module-bounded Global variables that are used for formatting purposes
_border_style_left = Border(left=Side(style='thin'))
_border_style_right = Border(right=Side(style='thin'))

# This function outlines the foundation for each of the tables. It defines the original row and column values
# and then calls functions in order to create tables for all specializations for CSE students 
def createCSESpecializationDocument(sheet, specializeCourses) -> str:
    
    sheet.merge_cells(start_row=1, start_column = 1, end_row=1, end_column=20)
    sheet.cell(row=1, column=1).value = "Optional specializations offered to CSE students"
    sheet.cell(row=1, column=1).font = Font(bold=True)

    row_value_start = 3
    start_col = 1
    end_col = 8 

    list_of_completed = ["Artificial Intelligence and Data Science Specialization Completed ", 
                        "Human-Computer Interaction Specialization Completed", 
                        "Game Programming Specialization Completed",
                        "Security and Privacy Specialization Completed", 
                        "Systems Software Development Specialization Completed"] 
    index_text = -1

    # Will return the next column to start at 
    dataScienceStatus = artificialAndDataScience(sheet, row_value_start, start_col, end_col, 4, specializeCourses)
    if dataScienceStatus == "Yes": index_text = 0

    start_col = end_col + 2
    end_col = start_col + 8 

    humanStatus = humanCompInteraction(sheet, row_value_start, start_col, end_col, 7, specializeCourses)
    if humanStatus == "Yes": index_text = 1

    start_col = end_col + 2
    end_col = start_col + 8

    gameStatus = gameProgramming(sheet, row_value_start, start_col, end_col, 7, specializeCourses)
    if gameStatus == "Yes": index_text = 2

    start_col = end_col + 2
    end_col = start_col + 8

    securityStatus = securityPrivacy(sheet, row_value_start, start_col, end_col, 5, specializeCourses)
    if securityStatus == "Yes": index_text = 3

    start_col = end_col + 2
    end_col = start_col + 8

    systemsStatus = systemsSoftware(sheet, row_value_start, start_col, end_col, 5, specializeCourses)
    if systemsStatus == "Yes": index_text = 4

    if index_text != -1:
        return list_of_completed[index_text] # Get the string associated with it 
    else:
        return "" # Return an empty string if it is still -1

# Outlines the artificial and data science table and displays all the required classes needed to take   
@staticmethod
def artificialAndDataScience(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:

    specializeCopy = specializeCourses.copy()
    
    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Artificial Intelligence and Data Science Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    # Create the objectives line first 
    required_starter = row_value_start + 1

    specializationMessage = "Requires 4 courses, at least 2 of them must be from Core Courses"
    nextRowStarter = specializationObjectives(sheet, required_starter, column_value_start, column_value_end, specializationMessage)


    valuesToPopAndInsert1 = {"Core Courses": 4, "Courses Satisfied": 4}
    required_course_one_data = ["CSE 351", "CSE 352", "CSE 353", "CSE 357"]

    row_limit_core = len(required_course_one_data)
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, nextRowStarter, column_value_start, column_value_end, row_limit_core, valuesToPopAndInsert1, required_course_one_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    valuesToPopAndInsert2 = {"Electives": 4, "Courses Satisfied": 4}
    required_course_two_data = ["CSE/ISE/EST 323", "CSE/ISE 332", "CSE/ISE 337", "CSE/MAT 371", "CSE 327, 354, 378", "CSE 390*, 391*, 392*, 393*, 394*", "CSE 487*", "CSE 495*, 496*"]

    # Same row_limit
    row_limit_electives = len(required_course_two_data)
    nextSectionRowStarter2 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, row_limit_electives, valuesToPopAndInsert2, required_course_two_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter2[1]

    # Create disclaimer for specialization 
    disclaimerMessage = "*Special topic or research project must be in Artificial Intelligence or Data Science."
    nextSectionRowStarter3 = specializationDisclaimer(sheet, nextSectionRowStarter2[0], column_value_start, column_value_end, disclaimerMessage)

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter3, column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus

# Outlines the human computer interaction table and displays all the required classes needed to take   
@staticmethod
def humanCompInteraction(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:
    
    specializeCopy = specializeCourses.copy()
    
    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Human-Computer Interaction Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    required_starter = row_value_start + 1

    specializationMessage = "Requires 4 courses, 2 electives, and 1 project"
    nextRowStarter = specializationObjectives(sheet, required_starter, column_value_start, column_value_end, specializationMessage)


    valuesToPopAndInsert1 = {"Core Courses": 4, "Courses Satisfied": 4}
    required_course_one_data = ["CSE/ISE/EST 323", "PSY 260", "CSE 328", "CSE/ISE 332", "CSE/ISE 333", "PSY 384"]

    row_limit_core = len(required_course_one_data)
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, nextRowStarter, column_value_start, column_value_end, row_limit_core, valuesToPopAndInsert1, required_course_one_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    valuesToPopAndInsert2 = {"Electives": 4, "Courses Satisfied": 4}
    required_course_two_data = ["CSE/ISE 332", "CSE/ISE 333", "CSE/ISE 334", "CSE/ISE 364", "CSE 327, 328, 336, 352, 366, 378", "CSE 390*, 391*, 392*, 393*, 394*", "PSY 366, 368, 369, 384"]

    # Same row_limit
    row_limit_electives = len(required_course_two_data)
    nextSectionRowStarter2 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, row_limit_electives, valuesToPopAndInsert2, required_course_two_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter2[1]


    # Same row_limit
    valuesToPopAndInsert3 = {"Project Requirement": 4, "Courses Satisfied": 4}
    required_course_data_three = ["CSE 487*", "CSE 488*", "CSE 495*, 496*"]

    row_limit_project = 3
    nextSectionRowStarter3 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter2[0], column_value_start, column_value_end, row_limit_project, valuesToPopAndInsert3, required_course_data_three, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter3[1]


    disclaimerMessage = "*Special topic or research project must be in Human-Computer Interaction"
    nextSectionRowStarter4 = specializationDisclaimer(sheet, nextSectionRowStarter3[0], column_value_start, column_value_end, disclaimerMessage)

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter4, column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus  

# Outlines the game programming table and displays all the required classes needed to take   
@staticmethod
def gameProgramming(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:

    specializeCopy = specializeCourses.copy()

    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Game Programming Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    required_starter = row_value_start + 1

    specializationMessage = "Requires 4 courses, 2 electives, and 1 project"
    nextRowStarter = specializationObjectives(sheet, required_starter, column_value_start, column_value_end, specializationMessage)

    valuesToPopAndInsert1 = {"Core Courses": 4, "Courses Satisfied": 4}
    required_course_one_data = ["CSE 306", "CSE 328", "CSE 380", "CSE 381"]

    row_limit_core = len(required_course_one_data)
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, nextRowStarter, column_value_start, column_value_end, row_limit_core, valuesToPopAndInsert1, required_course_one_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    valuesToPopAndInsert2 = {"Electives": 4, "Courses Satisfied": 4}
    required_course_two_data = ["CSE 327", "CSE/ISE 332", "CSE/ISE 334", "CSE/ISE 364", "CSE 355/AMS 345", "CSE 331, 352, 353, 376, 378"]

    # Same row_limit
    row_limit_electives = len(required_course_two_data)
    nextSectionRowStarter2 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, row_limit_electives, valuesToPopAndInsert2, required_course_two_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter2[1]

    # Same row_limit
    valuesToPopAndInsert3 = {"Project Requirement": 4, "Courses Satisfied": 4}
    required_course_data_three = ["CSE 487*", "CSE 488*", "CSE 495*, 496*"]

    row_limit_project = len(required_course_data_three)
    nextSectionRowStarter3 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter2[0], column_value_start, column_value_end, row_limit_project, valuesToPopAndInsert3, required_course_data_three, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter3[1]


    disclaimerMessage = "*Special topic or research project must be in Game Programming"
    nextSectionRowStarter4 = specializationDisclaimer(sheet, nextSectionRowStarter3[0], column_value_start, column_value_end, disclaimerMessage)

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter4, column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus  

# Outlines the security and privacy table and displays all the required classes needed to take   
@staticmethod
def securityPrivacy(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:
    
    # Might need to come back here to modify certain specifications, so we will have to see when we come back

    specializeCopy = specializeCourses.copy()
    
    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Security and Privacy Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    required_starter = row_value_start + 1

    specializationMessage = "Requires 2 core courses and 3 electives"
    nextRowStarter = specializationObjectives(sheet, required_starter, column_value_start, column_value_end, specializationMessage)

    valuesToPopAndInsert1 = {"Core Courses": 4, "Courses Satisfied": 4}
    required_course_one_data = ["CSE 331", "CSE 360, 361, 362, 363"]

    row_limit_core = len(required_course_one_data)
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, nextRowStarter, column_value_start, column_value_end, row_limit_core, valuesToPopAndInsert1, required_course_one_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    # Might need to put another disclaimer line for this specialization
    valuesToPopAndInsert2 = {"Electives": 4, "Courses Satisfied": 4}
    required_course_two_data = ["CSE 360", "CSE 361", "CSE 362", "CSE 363", "CSE 304 or CSE 307", "CSE 306 or CSE 356 or CSE 376", "CSE 390*, 391*, 392*, 392*, 394*", "CSE 487*, 495*, 496*"]

    # Need to put a securityElectivesDisclaimer

    electiveMessage = "Note: Does not include electives taken as a core course"
    electiveMessage2 = "Note: At most one course from each item may be used to satisfy specialization requirements"

    nextSectionRowStarter2 = internalDisclaimer(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, electiveMessage)

    nextSectionRowStarter3 = internalDisclaimer(sheet, nextSectionRowStarter2, column_value_start, column_value_end, electiveMessage2)

    # Same row_limit
    row_limit_electives = len(required_course_two_data)
    nextSectionRowStarter4 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter3, column_value_start, column_value_end, row_limit_electives, valuesToPopAndInsert2, required_course_two_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter4[1]

    disclaimerMessage = "*Special topic or research project must be in Security and Privacy"
    nextSectionRowStarter5 = specializationDisclaimer(sheet, nextSectionRowStarter4[0], column_value_start, column_value_end, disclaimerMessage)

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter5, column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus  

# Outlines the systems software table and displays all the required classes needed to take   
@staticmethod
def systemsSoftware(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:

    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, '': 0.00}

    specializeCopy = specializeCourses.copy()

    # Do some dictionary handling before we begin
    # We have to keep in mind that at most 2 courses may be drawn from CSE 331, CSE 360 - 363
    # Keep the values in the dictionary that have credit

    check_list = ["CSE 331", "CSE 360", "CSE 361", "CSE 362", "CSE 363"]
    safe_index = [False, False, False, False, False]

    at_most_counter = 0 
    for i in check_list:
        if i in specializeCopy.keys(): # If the value is also in the dictionary's set of keys
            # Check the credits earned for the course and if it is non-zero, then that key is safe 
            getGrade = specializeCopy[i].grade
            try:
                valueCorresponding = acceptable_grades[getGrade]
            except:
                valueCorresponding = 0.00
            if float(valueCorresponding) != 0.00 and at_most_counter < 2: # Then that class is safe
                indexOfSafe = check_list.index(i)
                safe_index[indexOfSafe] = True
                at_most_counter += 1
    
    # Now remove the classes that are not safe
    for index, classBool in enumerate(safe_index):
        if classBool == False and check_list[index] in specializeCopy.keys(): # If false and the class does exist 
            getClass = check_list[index]
            del specializeCopy[getClass] # Delete copy from dictionary
        

    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Security and Privacy Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    required_starter = row_value_start + 1

    specializationMessage = "Requires 5 of the following courses"
    nextRowStarter = specializationObjectives(sheet, required_starter, column_value_start, column_value_end, specializationMessage)

    electiveMessage = "Note: At most two of the courses may be drawn from CSE 331, CSE 360-363"
    nextRowStarter2 = internalDisclaimer(sheet, nextRowStarter, column_value_start, column_value_end, electiveMessage)

    valuesToPopAndInsert1 = {"Required Courses": 4, "Courses Satisfied": 4}
    required_course_one_data = ["CSE 331", "CSE 360, 361, 362, 363", "CSE 304", "CSE 306", "CSE/ISE 311", "CSE 356", "CSE 376", "CSE 390*, 391*, 392*, 393*, 394*"]

    row_limit_core = len(required_course_one_data)
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, nextRowStarter2, column_value_start, column_value_end, row_limit_core, valuesToPopAndInsert1, required_course_one_data, specializeCopy)
    total_courses_required_counter += nextSectionRowStarter1[1]

    disclaimerMessage = "*Special topic or research project must be in Systems Software Development"
    nextSectionRowStarter5 = specializationDisclaimer(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, disclaimerMessage)

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter5, column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus  

# This function is responsible for creating the table as well as inputting the required classes that are needed to satisfy
# a particular specialization. In addition, if a course that a student takes satisfies one of the required courses, then
# that course will be noted down next to the required course that was satisfied.
@staticmethod
def specializationTableInputterAlgorithm(sheet, row_start, column_start, column_end, row_limit, insert_dictionary, required_course_data, specializeCourses) -> list:

    class_sum_to_return = 0 

    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, '': 0.00}

    sheet.merge_cells(start_row=row_start, start_column=column_start, end_row=row_start, end_column=column_start+3) # For The First 3 columns
    sheet.merge_cells(start_row=row_start, start_column=column_start+4, end_row=row_start, end_column=column_end) # For the last 2 columns

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    initial_column = column_start 
    keys_in_dict = list(insert_dictionary.keys())
    while len(insert_dictionary) != 0:
        sheet.cell(row=row_start, column=initial_column).value = keys_in_dict[0]
        sheet.cell(row=row_start, column=initial_column).font = Font(size=10, name='Calibri', bold=True)
        valueOffset = insert_dictionary[keys_in_dict[0]]
        del insert_dictionary[keys_in_dict[0]]
        keys_in_dict.pop(0)
        initial_column += valueOffset
    
    initial_column = column_start
    while initial_column != column_end + 1:
        sheet.cell(row=row_start, column=initial_column).border = all_sides_border
        initial_column += 1
    
    row_lower = row_start + 1
    row_max = row_lower + row_limit

    # Responsible for injecting the preassumed data 
    keys_of_courses = list(required_course_data)

    while row_lower != row_max: # Injects some presumed data 
        sheet.merge_cells(start_row=row_lower, start_column=column_start, end_row=row_lower, end_column=column_start+3)
        sheet.merge_cells(start_row=row_lower, start_column=column_start+4, end_row=row_lower, end_column=column_end)
        try:
            keyFirst = keys_of_courses[0]
            courseName = keyFirst
            sheet.cell(row=row_lower, column=column_start).value = courseName
        except: pass
        sheet.cell(row=row_lower, column=column_start).font = Font(size=10)
        sheet.cell(row=row_lower, column=column_start).border = _border_style_left # Border to Lower Division
        sheet.cell(row=row_lower, column=column_start+3).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+4).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+5).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+6).border = _border_style_left
        sheet.cell(row=row_lower, column=column_end-1).border = _border_style_left
        sheet.cell(row=row_lower, column=column_end).border = _border_style_right
        try:
            keys_of_courses.pop(0)
        except:
            pass
        row_lower += 1
    
    # Before returning, we now check to see if the course exists in our specializeCourses dictionary 
    getListOfSortedKeys = list(sorted(specializeCourses.keys()))
    lengthOfList = len(getListOfSortedKeys)

    searching_start_row = row_start + 1
    row_offset = 1


    while lengthOfList != 0:

        getFirstIndex = getListOfSortedKeys[0]

        if searching_start_row >= row_max: # Not found, so we pop this key from the list and continue 

            searching_start_row = row_start + 1
            row_offset = 1
            getListOfSortedKeys.pop(0)
            lengthOfList -= 1
            continue # So we can start looking at the next element in the list right away 
        
        
        cellValue = sheet.cell(row=searching_start_row, column=column_start).value

        boolCSEISE, boolCSEAMS, boolCSEMAT, boolCSEISEEST = False, False, False, False 

        if "CSE/ISE" in getFirstIndex:
            if getFirstIndex[0:3] in cellValue and getFirstIndex[4:7] in cellValue:
                boolCSEISE = True
        
        elif "CSE/MAT" in getFirstIndex:
            if getFirstIndex[0:3] in cellValue and getFirstIndex[4:7] in cellValue:
                boolCSEMAT = True 
  
        elif "CSE 355/AMS 345" in getFirstIndex:
            if getFirstIndex[0:3] in cellValue and getFirstIndex[8:11] in cellValue:
                boolCSEAMS = True 
        
        elif "CSE/ISE/EST 323" in getFirstIndex:
            if getFirstIndex[0:3] in cellValue and getFirstIndex[4:7] in cellValue and getFirstIndex[8:11]:
                boolCSEISEEST = True
            

        # First check the description if it is in the cellValue 
        if getFirstIndex[0:3] in cellValue or boolCSEISE == True or boolCSEMAT == True or boolCSEAMS == True or boolCSEISEEST == True: # Found the class, now check if the grade is of acceptable value 
            
            acceptDiff = False 

            index1 = 4 if boolCSEISE == False or boolCSEMAT == False else 8 
            index2 = 7 if boolCSEISE == False or boolCSEMAT == False else 11

            if boolCSEAMS == True and getFirstIndex[4:7] in cellValue and getFirstIndex[12:15] in cellValue: acceptDiff = True

            index1 = index1 if boolCSEISEEST == False else 12
            index2 = index2 if boolCSEISEEST == False else 15

            if getFirstIndex[index1:index2] in cellValue or acceptDiff == True:
                
                getGrade = specializeCourses[getFirstIndex].grade
                try:
                    valueCorresponding = acceptable_grades[getGrade]
                except:
                    valueCorresponding = 0.00
                if float(valueCorresponding) == 0.00: # We don't count the classes that don't have credit, but we don't delete it from the dictionary just yet 
                    searching_start_row = row_start + 1 # Reset back to original spot and use a new class 
                    row_offset = 1
                    getListOfSortedKeys.pop(0)
                    lengthOfList -= 1  
                else:
                    class_sum_to_return += 1
                    currentCellValue = sheet.cell(row=searching_start_row, column=column_start+4).value
                    if currentCellValue == None: # Empty cell 
                        sheet.cell(row=searching_start_row, column=column_start+4).value = f'{getFirstIndex}'
                        del specializeCourses[getFirstIndex] # Delete the key from the dictionary so it wont be used again   
                    else:
                        newCellValue = ""
                        if currentCellValue[0:3] == getFirstIndex[0:3]: # Same class Description
                            newCellValue = currentCellValue + ", " + f'{getFirstIndex[4:7]}' # We just put the class number
                        else: # Different class description
                            newCellValue = currentCellValue + ", " + f'{getFirstIndex}'
                        sheet.cell(row=searching_start_row, column=column_start+4).value = str(newCellValue)
                        del specializeCourses[getFirstIndex] # Delete the key from the dictionary so it wont be used again   

                    searching_start_row = row_start + 1 # Reset back to original spot 
                    row_offset = 1
                    getListOfSortedKeys.pop(0)
                    lengthOfList -= 1  
             
            else:
                searching_start_row += 1
                row_offset += 1              
        else:
            searching_start_row += 1
            row_offset += 1


    return [row_max, class_sum_to_return] # Return a list


@staticmethod
def specializationObjectives(sheet, row_value, column_value_start, column_value_end, specializationMessage) -> int:
    
    # First merge the cells 
    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_value, column=start_col_merge).border = all_sides_border
        start_col_merge += 1

    sheet.merge_cells(start_row=row_value, start_column=column_value_start, end_row=row_value, end_column=column_value_end)
    sheet.cell(row=row_value, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_value, column=column_value_start).value = specializationMessage

    return row_value + 1 # Return the next row starting value 

# This function outlines any special message (i.e. special project or research) and displays it on the table as a disclaimer
@staticmethod
def specializationDisclaimer(sheet, row_value, column_value_start, column_value_end, disclaimerMessage) -> int:
    # First merge the cells 
    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_value, column=start_col_merge).border = all_sides_border
        start_col_merge += 1

    sheet.merge_cells(start_row=row_value, start_column=column_value_start, end_row=row_value, end_column=column_value_end)
    sheet.cell(row=row_value, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_value, column=column_value_start).value = disclaimerMessage

    return row_value + 1 # Return the next row starting value 

# This function outlines another kind of special message, one that pertains more to a certain course amount requirement 
# For example, at most two courses are needed 
@staticmethod
def internalDisclaimer(sheet, row_value, column_value_start, column_value_end, electiveDisclaimerMessage) -> int:

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    start_col_merge = column_value_start
    
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_value, column=start_col_merge).border = all_sides_border
        start_col_merge += 1

    sheet.merge_cells(start_row=row_value, start_column=column_value_start, end_row=row_value, end_column=column_value_end)
    sheet.cell(row=row_value, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_value, column=column_value_start).value = electiveDisclaimerMessage

    return row_value + 1 # Return the next row starting value 

# This function displays a message as to whether or not the corresponding specialization was satisfied or not.
# The message will display specialization completed if the student has completed all the required courses. Otherwise, it will display not completed
@staticmethod
def specalizationCompleteQuestionLine(sheet, row_value, column_value_start, column_value_end, row_limit, total_class_count, class_amount_needed) -> str:

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    row_lower = row_value + 1
    row_max = row_lower + row_limit

    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_max, column=start_col_merge).border = all_sides_border
        start_col_merge += 1

    sheet.merge_cells(start_row=row_max, start_column=column_value_start, end_row=row_max, end_column=column_value_end)
    sheet.cell(row=row_max, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)

    if total_class_count >= class_amount_needed: # Means specialization completed
        sheet.cell(row=row_max, column=column_value_start).value = f'Status: Specialization Completed ({total_class_count} of {class_amount_needed})'
        return "Yes"
    else:
        sheet.cell(row=row_max, column=column_value_start).value = f'Status: Specialization Incomplete ({total_class_count} of {class_amount_needed})'
        return "No"