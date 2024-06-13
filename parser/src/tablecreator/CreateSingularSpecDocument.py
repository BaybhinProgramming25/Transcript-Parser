"""
The purpose of this is to create the singular specialization document for 
The students who have a specialization declared

If a student did not declare a specialization, then 
This table will not appear in the same sheet as the other three tables 

"""


from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side


_border_style_left = Border(left=Side(style='thin'))
_border_style_right = Border(right=Side(style='thin'))

# The purpose of this function is to first determine the specialization declared by the student, if any. 
# Once a specialization has been found, when then create the associated table with it via a function call 
def createSingularSpecDocument(sheet, row_value_start, column_value_start, column_value_end, studentInformation, specalizeISECourses, specializeCSECourses, technicalCSECourses):

    if (column_value_end - column_value_start + 1) % 3 != 0:
        column_value_start += 2 # Shift by 2 
        column_value_end += 3

    column_counter = column_value_start

    if studentInformation['Major'] == "CSE":
        specializationName = ""
        if studentInformation['Spec'] == "Artificial":
            specializationName = "Artificial Intelligence and Data Science Specialization Requirements"
        elif studentInformation['Spec'] == "Interaction":
            specializationName = "Human-Computer Interaction Specialization Requirements"
        elif studentInformation['Spec'] == "Game":
            specializationName = "Game Programming Specialization Requirements"
        elif studentInformation['Spec'] == "Security":
            specializationName = "Security and Privacy Specialization Requirements"
        elif studentInformation['Spec'] == "System":
            specializationName = "Systems Software Development Specialization Requirements"
        
        if specializationName != "":
            sheet.cell(row=row_value_start, column=column_value_start).value = specializationName
        else:
            sheet.merge_cells(start_row=row_value_start, start_column = column_value_start, end_row=row_value_start, end_column=column_value_end)
            sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
            sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
            sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')
            sheet.cell(row=row_value_start, column=column_value_start).value = "Please Declare A Specialization To View The Specialization Table"
            return # Return, as we have nothing left to do besides displaying the message       
    elif studentInformation['Major'] == "ISE":
        specializationName = ""
        if studentInformation['Spec'] == "Network":
            specializationName = "Systems and Network Administration Specialization Requirements"
        elif studentInformation['Spec'] == "Health":
            specializationName = "Health Informatics Specialization Requirements"
        elif studentInformation['Spec'] == "Economics":
            specializationName = "Business and Economics Specialization Requirements"
        elif studentInformation['Spec'] == "Technological":
            specializationName = "Technological Systems Management Specialization Requirements"
        elif studentInformation['Spec'] == "Financial":
            specializationName = "Financial Information Specialization Requirements"
        
        if specializationName != "":
            sheet.cell(row=row_value_start, column=column_value_start).value = specializationName
        else:
            sheet.merge_cells(start_row=row_value_start, start_column = column_value_start, end_row=row_value_start, end_column=column_value_end)
            sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
            sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
            sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')
            sheet.cell(row=row_value_start, column=column_value_start).value = "Please declare a specialization to view specialization table"
            return # Return, as we have nothing left to do besides displaying the message  
        
    sheet.merge_cells(start_row=row_value_start, start_column = column_value_start, end_row=row_value_start, end_column=column_value_end)
    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')


    # After formatting, we begin the actual insertion of tables 

    if studentInformation['Major'] == "ISE":
        if studentInformation['Spec'] == "Network":
            createNetworkTable(sheet, row_value_start, column_value_start, column_value_end, specalizeISECourses, studentInformation) 
        elif studentInformation['Spec'] == "Health":
            createHealthTable(sheet, row_value_start, column_value_start, column_value_end, specalizeISECourses, studentInformation)
        elif studentInformation['Spec'] == "Economics":
            createEconomicsTable(sheet, row_value_start, column_value_start, column_value_end, specalizeISECourses, studentInformation)
        elif studentInformation['Spec'] == "Technological":
            createTechnologicalTable(sheet, row_value_start, column_value_start, column_value_end, specalizeISECourses, studentInformation)
        elif studentInformation['Spec'] == "Financial":
            createFinancialTable(sheet, row_value_start, column_value_start, column_value_end, specalizeISECourses, studentInformation)
    elif studentInformation['Major'] == "CSE":
        specializeCSECourses.update(technicalCSECourses) # Combine them into one dictionary
        if studentInformation['Spec'] == "Artificial":
            createArtificialTable(sheet, row_value_start, column_value_start, column_value_end, specializeCSECourses, studentInformation)
        elif studentInformation['Spec'] == "Interaction":
            createInteractionTable(sheet, row_value_start, column_value_start, column_value_end, specializeCSECourses, studentInformation)
        elif studentInformation['Spec'] == "Game":
            createGameTable(sheet, row_value_start, column_value_start, column_value_end, specializeCSECourses, studentInformation)
        elif studentInformation['Spec'] == "Security":
            createSecurityTable(sheet, row_value_start, column_value_start, column_value_end, specializeCSECourses, studentInformation)
        elif studentInformation['Spec'] == "System":
            createSystemTable(sheet, row_value_start, column_value_start, column_value_end, specializeCSECourses, studentInformation)


##################################### ISE COURSES TABLES BEGIN HERE ###################################################

# Function defines network table 
@staticmethod
def createNetworkTable(sheet, row_starter, column_value_start, column_value_end, specalizeISECourses, studentInformation):
    
    row_begin_insert = row_starter + 1
    row_limit = 1 

    valuesToPopAndInsert = {"Required Course #1": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE/ISE 311"]


    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_begin_insert, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)
    
    valuesToPopAndInsert = {"Required Course #2": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["ISE 321"]

    returnNextSectionRowStarter2 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)

    valuesToPopAndInsert = {"ONE of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2} 
    required_course = ["ISE 331", "CSE 331"] # 2 Different courses 

    row_limit = 2
    returnNextSectionRowStarter3 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter2, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)

    valuesToPopAndInsert = {"ONE of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2} 
    required_course = ["CSE/ISE 337"]
    
    row_limit = 1
    returnNextSectionRowStarter4 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter3, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)

    valuesToPopAndInsert = {"ONE of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2} 
    required_course = ["ISE 475", "ISE 488"] # 2 Different courses

    row_limit = 2
    returnNextSectionRowStarter5 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter4, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)

    valuesToPopAndInsert = {"ONE of the Required Courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2} 
    required_course = ["ESE 442", "CSE 370", "EST 393", "BUS 393", "BUS 346", "AMS 341"]

    row_limit = 6
    excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter5, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)

# Function defines health table 
@staticmethod
def createHealthTable(sheet, row_starter, column_value_start, column_value_end, specalizeISECourses, studentInformation):

    row_begin_insert = row_starter + 1
    row_limit = 1 

    valuesToPopAndInsert = {"Required Course #1": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["HAN 200"]

    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_begin_insert, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)

    valuesToPopAndInsert = {"ONE of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["BIO 202", "BIO 203"]

    row_limit = 2
    returnNextSectionRowStarter2 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)

    valuesToPopAndInsert = {"FOUR of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2} 
    required_course = ["BCP 405", "BME 205", "CSE 377", "ECO 327", "HAN 202", "PSY 103"]

    row_limit = 6
    excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter2, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 4, studentInformation)
    
# Function defines economics table 
@staticmethod
def createEconomicsTable(sheet, row_starter, column_value_start, column_value_end, specalizeISECourses, studentInformation):
    row_begin_insert = row_starter + 1
    row_limit = 1 

    valuesToPopAndInsert = {"Required Course #1": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["ECO 108"]

    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_begin_insert, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)
    
    valuesToPopAndInsert = {"Required Course #2": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["ACC 210"]

    returnNextSectionRowStarter2 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 1, studentInformation)

    valuesToPopAndInsert = {"TWO of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2} 
    required_course = ["ACC 214", "ESE 201", "BUS 215", "BUS 220", "BUS 294"]

    row_limit = 5
    returnNextSectionRowStarter3 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter2, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 2, studentInformation)

    valuesToPopAndInsert = {"FOUR of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2} 
    required_course = ["BUS 330", "BUS 346", "BUS 348", "BUS 355", "BUS 356", "EST 305", "EST 320", "EST 325", "EST 364", "EST 392", "EST 393", "ECO 326", "ECO 348", "ECO 389", "POL 319", "POL 359", "SOC 381"]

    row_limit = len(required_course)
    excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter3, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 4, studentInformation)

# Function defines technological table 
@staticmethod
def createTechnologicalTable(sheet, row_starter, column_value_start, column_value_end, specalizeISECourses, studentInformation):

    row_begin_insert = row_starter + 1

    valuesToPopAndInsert = {"FOUR of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["EST 201", "EST 202", "EST 391", "EST 392", "EST 393"]

    row_limit = len(required_course)
    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_begin_insert, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 4, studentInformation)

    valuesToPopAndInsert = {"TWO of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["ISE 340/EST 310", "EST 320", "EST 323", "CSE/ISE/EST 323", "EST 326", "EST 327", "EST 364"]

    row_limit = len(required_course)
    excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 2, studentInformation)

# Function defines financial table 
@staticmethod
def createFinancialTable(sheet, row_starter, column_value_start, column_value_end, specalizeISECourses, studentInformation):

    row_begin_insert = row_starter + 1

    valuesToPopAndInsert = {"TWO of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 215", "AMS 315", "AMS 318"]

    row_limit = len(required_course)
    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_begin_insert, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 2, studentInformation)

    valuesToPopAndInsert = {"FOUR of the Required courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["ACC 210", "AMS 311", "AMS 316", "AMS 320", "AMS 341", "AMS 394", "AMS 441", "BUS 330", "BUS 331", "BUS 355", "BUS 356", "CSE/ISE/EST 323", "ISE 331"]

    row_limit = len(required_course)
    excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specalizeISECourses, 4, studentInformation)



################################ CSE COURSES TABLE BEGINS HERE #################################################

# Function defines artificial table 
@staticmethod
def createArtificialTable(sheet, row_starter, column_value_start, column_value_end, specializeCSECourses, studentInformation):

    row_begin_insert = row_starter + 1

    insertMessage = "Requires 4 courses, at least 2 of them must be from Core Courses"
    row_after_message = specializationObjectives(sheet, row_begin_insert, column_value_start, column_value_end, insertMessage)

    valuesToPopAndInsert = {"Core Courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 351", "CSE 352", "CSE 353", "CSE 357"]

    row_limit = len(required_course)
    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_after_message, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    valuesToPopAndInsert = {"Electives": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE/ISE/EST 323", "CSE 327", "CSE/ISE 332", "CSE/ISE 337", "CSE 354", "CSE/MAT 371", "CSE 378", "CSE 390*", "CSE 391*", "CSE 392*", "CSE 393*", "CSE 394*", "CSE 487*", "CSE 495*", "CSE 496*"]

    row_limit = len(required_course)
    returnNextSectionRowStarter2 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    insertMessage = "*Special topic or research project must be in Artificial Intelligence or Data Science."
    specializationObjectives(sheet, returnNextSectionRowStarter2, column_value_start, column_value_end, insertMessage)

# Function defines interaction table 
@staticmethod
def createInteractionTable(sheet, row_starter, column_value_start, column_value_end, specializeCSECourses, studentInformation):

    row_begin_insert = row_starter + 1

    insertMessage = "Requires 4 courses, 2 electives, and 1 project"
    row_after_message = specializationObjectives(sheet, row_begin_insert, column_value_start, column_value_end, insertMessage)

    valuesToPopAndInsert = {"Core Courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE/ISE/EST 323", "PSY 260", "CSE 328", "CSE/ISE 332", "CSE/ISE 333", "PSY 384"]

    row_limit = len(required_course)
    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_after_message, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    valuesToPopAndInsert = {"Electives": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 327", "CSE 328", "CSE/ISE 332", "CSE/ISE 333", "CSE/ISE 334", "CSE 336", "CSE 352", "CSE/ISE 364", "CSE 366", "CSE 378", "CSE 390*", "CSE 391*", "CSE 392*", "CSE 393*", "CSE 394*", "PSY 366", "PSY 368", "PSY 369", "PSY 384"]

    row_limit = len(required_course)
    returnNextSectionRowStarter2 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    valuesToPopAndInsert = {"Project Requirement": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 487*", "CSE 488*", "CSE 495*", "CSE 496*"]

    row_limit = len(required_course)
    returnNextSectionRowStarter3 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter2, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    insertMessage = "*Special topic or research project must be in Human-Computer Interaction"
    specializationObjectives(sheet, returnNextSectionRowStarter3, column_value_start, column_value_end, insertMessage)

# Function defines game table 
@staticmethod
def createGameTable(sheet, row_starter, column_value_start, column_value_end, specializeCSECourses, studentInformation):

    row_begin_insert = row_starter + 1

    insertMessage = "Requires 4 courses, 2 electives, and 1 project"
    row_after_message = specializationObjectives(sheet, row_begin_insert, column_value_start, column_value_end, insertMessage)

    valuesToPopAndInsert = {"Core Courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 306", "CSE 328", "CSE 380", "CSE 381"]

    row_limit = len(required_course)
    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_after_message, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    valuesToPopAndInsert = {"Electives": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 327", "CSE 331", "CSE/ISE 332", "CSE/ISE 334", "CSE 352", "CSE 353", "CSE 355/AMS 345", "CSE/ISE 364", "CSE 376", "CSE 378"]

    row_limit = len(required_course)
    returnNextSectionRowStarter2 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    valuesToPopAndInsert = {"Project Requirement": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 487*", "CSE 488*", "CSE 495*", "CSE 496*"]

    row_limit = len(required_course)
    returnNextSectionRowStarter3 = excelSpecInputterAlgorithm(sheet, returnNextSectionRowStarter2, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    insertMessage = "*Special topic or research project must be in Game Programming"
    specializationObjectives(sheet, returnNextSectionRowStarter3, column_value_start, column_value_end, insertMessage)

# Function defines security table 
@staticmethod
def createSecurityTable(sheet, row_starter, column_value_start, column_value_end, specializeCSECourses, studentInformation):

    row_begin_insert = row_starter + 1

    insertMessage = "Requires 2 core courses and 3 electives"
    row_after_message = specializationObjectives(sheet, row_begin_insert, column_value_start, column_value_end, insertMessage)

    valuesToPopAndInsert = {"Core Courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 331", "CSE 360", "CSE 361", "CSE 362", "CSE 363"]

    row_limit = len(required_course)
    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_after_message, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    insertMessage = "Note: Does not include electives taken as a core course"
    row_after_message_second = specializationObjectives(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, insertMessage)

    insertMessage = "Note: At most one course from each item may be used to satisfy specialization requirements"
    row_after_message_third = specializationObjectives(sheet, row_after_message_second, column_value_start, column_value_end, insertMessage)

    valuesToPopAndInsert = {"Electives": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 360", "CSE 361", "CSE 362", "CSE 363", "CSE 304", "CSE 307", "CSE 306", "CSE 356", "CSE 376", "CSE 390*", "CSE 391*", "CSE 392*", "CSE 393*", "CSE 394*", "CSE 487*", "CSE 488*", "CSE 495*", "CSE 496*"]

    row_limit = len(required_course)
    returnNextSectionRowStarter2 = excelSpecInputterAlgorithm(sheet, row_after_message_third, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    insertMessage = "*Special topic or research project must be in Security and Privacy"
    specializationObjectives(sheet, returnNextSectionRowStarter2, column_value_start, column_value_end, insertMessage)

# Function defines system table 
@staticmethod
def createSystemTable(sheet, row_starter, column_value_start, column_value_end, specializeCSECourses, studentInformation):

    row_begin_insert = row_starter + 1

    insertMessage = "Requires 5 of the following courses"
    row_after_message = specializationObjectives(sheet, row_begin_insert, column_value_start, column_value_end, insertMessage)

    insertMessage = "Note: At most two of the courses may be drawn from CSE 331, CSE 360-363"
    row_after_message_second = specializationObjectives(sheet, row_after_message, column_value_start, column_value_end, insertMessage)

    valuesToPopAndInsert = {"Required Courses": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}
    required_course = ["CSE 331", "CSE 360", "CSE 361", "CSE 362", "CSE 363", "CSE 304", "CSE 306", "CSE/ISE 311", "CSE 356", "CSE 376", "CSE 390*", "CSE 391*", "CSE 392*", "CSE 393*", "CSE 394*"]

    row_limit = len(required_course)
    returnNextSectionRowStarter = excelSpecInputterAlgorithm(sheet, row_after_message_second, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, required_course, specializeCSECourses, 0, studentInformation)

    insertMessage = "*Special topic or research project must be in Systems Software Development"
    specializationObjectives(sheet, returnNextSectionRowStarter, column_value_start, column_value_end, insertMessage)


# This function makes the table and inputs the specialization information into the table 
@staticmethod
def excelSpecInputterAlgorithm(sheet, row_start, column_start, column_end, row_limit, insert_dictionary, class_data, category_dictionary, number_of_classes, studentInformation) -> int:

    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, 'F': 0.00, 'IF': 0.00, 'Q': 0.00}

    sheet.merge_cells(start_row=row_start, start_column=column_start, end_row=row_start, end_column=column_start+2) # For Lower Division Courses
    sheet.merge_cells(start_row=row_start, start_column=column_end-1, end_row=row_start, end_column=column_end) # For Comments

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    initial_column = column_start 
    keys_in_dict = list(insert_dictionary.keys())
    while len(insert_dictionary) != 0:
        sheet.cell(row=row_start, column=initial_column).value = keys_in_dict[0]
        sheet.cell(row=row_start, column=initial_column).font = Font(size=10, name='Calibri', bold=True)
        valueOffset = insert_dictionary[keys_in_dict[0]]
        del insert_dictionary[keys_in_dict[0]]
        keys_in_dict.pop(0)
        initial_column += valueOffset
    
    initial_column = column_start
    while initial_column != column_end + 1:
        sheet.cell(row=row_start, column=initial_column).border = all_sides_border
        initial_column += 1
    
    row_lower = row_start + 1
    row_max = row_lower + row_limit

    keys_of_courses = []
    if len(class_data) == 0: # Means no preassumed data
        keys_of_courses = list(category_dictionary.keys())
    else: # There is preassumed data 
        keys_of_courses = list(class_data)


    while row_lower != row_max: # Injects some presumed data 
        sheet.merge_cells(start_row=row_lower, start_column=column_start, end_row=row_lower, end_column=column_start+2)
        sheet.merge_cells(start_row=row_lower, start_column=column_end-1, end_row=row_lower, end_column=column_end)
        try:
            keyFirst = keys_of_courses[0]
            courseName = keyFirst
            sheet.cell(row=row_lower, column=column_start).value = courseName
        except: pass
        sheet.cell(row=row_lower, column=column_start).font = Font(size=10)
        sheet.cell(row=row_lower, column=column_start).border = _border_style_left # Border to Lower Division
        sheet.cell(row=row_lower, column=column_start+3).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+4).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+5).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+6).border = _border_style_left
        sheet.cell(row=row_lower, column=column_end-1).border = _border_style_left
        sheet.cell(row=row_lower, column=column_end).border = _border_style_right
        try:
            keys_of_courses.pop(0)
        except:
            pass
        row_lower += 1
    
    start_col_merge = column_start
    while start_col_merge != column_end + 1:
        sheet.cell(row=row_lower, column=start_col_merge).border = all_sides_border
        start_col_merge += 1
    
    # Before returning, we now check to see if the course exists in our specializeCourses dictionary 
    getListOfSortedKeys = list(category_dictionary.keys())
    lengthOfList = len(getListOfSortedKeys)
    creditsCounter = 0
    creditsValue = 0
    section_satisfied = 0

    searching_start_row = row_start + 1
    row_offset = 1

    while lengthOfList != 0:

        getFirstIndex = getListOfSortedKeys[0]

        if searching_start_row >= row_max: # Not found, so we pop this key from the list and continue 

            searching_start_row = row_start + 1
            row_offset = 1
            getListOfSortedKeys.pop(0)
            lengthOfList -= 1
            continue # So we can start looking at the next element in the list right away 
        
        
        cellValue = sheet.cell(row=searching_start_row, column=column_start).value
        
        # No need to check the case for alternative classes, as they are statically typed 

        if getFirstIndex in cellValue: # Found the element  
                
                insertion_start = column_start + 3

                # Then we need to check and see if we have to replace the cell value 
                # We are only concerned for the GPA in a set number of classes 
                
                while insertion_start != column_end:

                    getColumnDescription = str(sheet.cell(row=searching_start_row-row_offset, column=insertion_start).value)
                    getColumnDescription = getColumnDescription.lower()
                    classObject = category_dictionary[getFirstIndex]
                    field_value = getattr(classObject, getColumnDescription)
                    try: 
                        sheet.cell(row=searching_start_row, column=insertion_start).value = float(field_value) # Put the grade of the object
                        if getColumnDescription == "credits":
                            getGradeLetter = str(sheet.cell(row=searching_start_row, column=insertion_start-1).value)
                            if getGradeLetter == 'IF':
                                getGradeLetter = 'F'
                            if getGradeLetter != 'XFER' and getGradeLetter != 'P' and getGradeLetter != 'S' and getGradeLetter != 'W' and getGradeLetter != 'U' and getGradeLetter != 'NC' and getGradeLetter != 'I' and getGradeLetter != '':
                                creditsCounter += float(field_value)
                                if float(field_value) > 0.00: # Means the student earned credit for that course 
                                    section_satisfied += 1
                                # We are also gonna wanna get the grade at the previous cell
                                getPoints = acceptable_grades[getGradeLetter]
                                letterAndCreditProduct = float(getPoints) * float(field_value)
                                creditsValue += letterAndCreditProduct
                    except: 
                            sheet.cell(row=searching_start_row, column=insertion_start).value = str(field_value) # Put the grade of the object 
                    sheet.cell(row=searching_start_row, column=insertion_start).alignment = Alignment(horizontal='left')
                    insertion_start += 1

                searching_start_row = row_start + 1 # Reset back to original spot 
                row_offset = 1
                getListOfSortedKeys.pop(0)
                lengthOfList -= 1         
        else: # If we didnt find the element, we keep searching but using the same element
                searching_start_row += 1 # Increment by 1 
                row_offset += 1    


    if creditsCounter != 0: totalGPA = round(float(creditsValue/creditsCounter), 3)
    else: totalGPA = round(float(0), 3)

    classSatisfied = "Yes" if section_satisfied >= number_of_classes else "No"

    booleanSatisfied = ""
    showClasses = ""
    totalGPALine = f'{totalGPA},' if studentInformation['Major'] == "ISE" else f'{totalGPA}'
    if studentInformation['Major'] == "ISE": # We don't include the boolean satisfied
        showClasses = f'({number_of_classes} classes)' 
        booleanSatisfied = f'Section Satisfied?: {classSatisfied}'


    sheet.merge_cells(start_row=row_max, start_column=column_start, end_row=row_max, end_column=column_end)
    sheet.cell(row=row_max, column=column_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_max, column=column_start).value = f'Category GPA (excluding P, S, XFER, etc courses) {showClasses}: {totalGPALine} {booleanSatisfied}'
    return row_max+1

# This function displays a specialization message on certain caveats about the specialization (i.e. special research or project, etc)
@staticmethod
def specializationObjectives(sheet, row_value, column_value_start, column_value_end, specializationMessage) -> int:
    
    # First merge the cells 
    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_value, column=start_col_merge).border = all_sides_border
        start_col_merge += 1

    sheet.merge_cells(start_row=row_value, start_column=column_value_start, end_row=row_value, end_column=column_value_end)
    sheet.cell(row=row_value, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_value, column=column_value_start).value = specializationMessage

    return row_value + 1 # Return the next row starting value 