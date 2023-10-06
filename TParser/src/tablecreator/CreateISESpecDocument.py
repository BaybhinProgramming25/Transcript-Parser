"""
This module is responsible for outlining all the possible
specializations that are offered to ISE students. A specialization
is required for all ISE students.

These specializations will be on a separate sheet, apart from
the necessary information that is displayed on the first sheet 

"""


from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

# Module-bounded Global variables that are used for formatting purposes
_border_style_left = Border(left=Side(style='thin'))
_border_style_right = Border(right=Side(style='thin'))

# This function outlines the foundation for each of the tables. It defines the original row and column values
# and then calls functions in order to create tables for all specializations for ISE students 
def createISESpecializationDocument(sheet, specializeCourses) -> str:
    
    sheet.merge_cells(start_row=1, start_column = 1, end_row=1, end_column=20)
    sheet.cell(row=1, column=1).value = "ISE Students are required to complete ONE of the following specializations (Note: A course will be noted down if and only if the student earns credit for that course)"
    sheet.cell(row=1, column=1).font = Font(bold=True)

    row_value_start = 3
    start_col = 1
    end_col = 8 

    list_of_completed = ["Business and Economics Specialization Completed", 
                        "System and Network Administration Specialization Completed", 
                        "Technological Systems Management Specialization Completed",
                        "Financial Information Systems Specialization Completed", 
                        "Health Informatics Specialization Completed"] 
    index_text = -1

    # Will return the next column to start at 
    businessStatus = createBussinessAndEconomicsTable(sheet, row_value_start, start_col, end_col, 6, specializeCourses)
    if businessStatus == "Yes": index_text = 0

    start_col = end_col + 2
    end_col = start_col + 8 

    systemsStatus = createSystemsAndNetworkTable(sheet, row_value_start, start_col, end_col, 6, specializeCourses)
    if systemsStatus == "Yes": index_text = 1

    start_col = end_col + 2
    end_col = start_col + 8

    techStatus = techAndSystemsTable(sheet, row_value_start, start_col, end_col, 6, specializeCourses)
    if techStatus == "Yes": index_text = 2

    start_col = end_col + 2
    end_col = start_col + 8

    financeStatus = financeSystemsTable(sheet, row_value_start, start_col, end_col, 6, specializeCourses)
    if financeStatus == "Yes": index_text = 3

    start_col = end_col + 2
    end_col = start_col + 8

    healthStatus = healthInformaticsTable(sheet, row_value_start, start_col, end_col, 6, specializeCourses)
    if healthStatus == "Yes": index_text = 4

    if index_text != -1:
        return list_of_completed[index_text] # Get the string associated with it 
    else:
        return "" # Return an empty string if it is still -1

# Outlines the business and economics table and displays all the required classes needed to take   
@staticmethod
def createBussinessAndEconomicsTable(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:

    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Business and Economics Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    valuesToPopAndInsert1 = {"Required Course #1": 4, "Courses Satisfied": 4}
    required_course_one_data = ["ECO 108"]

    required_starter = row_value_start + 1
    row_limit_one = 1
    
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, required_starter, column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert1, required_course_one_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    valuesToPopAndInsert2 = {"Required Course #2": 4, "Courses Satisfied": 4}
    required_course_two_data = ["ACC 210"]

    # Same row_limit
    nextSectionRowStarter2 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert2, required_course_two_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter2[1]


    valuesToPopAndInsert3 = {"2 of the following courses": 4, "Courses Satisfied": 4}
    required_course_data_three = ["ACC 214", "ESE 201", "BUS 215, 220, 294"]

    # Different row_limit

    row_limit_three = 3
    nextSectionRowStarter3 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter2[0], column_value_start, column_value_end, row_limit_three, valuesToPopAndInsert3, required_course_data_three, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter3[1]


    valuesToPopAndInsert4 = {"2 of the following courses": 4, "Courses Satisfied": 4}
    required_course_data_four = ["BUS 330, 346, 348, 355, 356", "EST 305, 320, 325, 364, 392, 393", "ECO 326, 348, 389", "POL 319, 359", "SOC 381"]

    # Different row_limit 
    row_limit_four = 5
    nextSectionRowStarter4 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter3[0], column_value_start, column_value_end, row_limit_four, valuesToPopAndInsert4, required_course_data_four, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter4[1]

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter4[0], column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus

# Outlines the systems and network table and displays all the required classes needed to take   
@staticmethod
def createSystemsAndNetworkTable(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:

    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Systems and Network Administration Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    valuesToPopAndInsert1 = {"Required Course #1": 4, "Courses Satisfied": 4}
    required_course_one_data = ["CSE/ISE 311"]

    required_starter = row_value_start + 1
    row_limit_one = 1
    
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, required_starter, column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert1, required_course_one_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    valuesToPopAndInsert2 = {"Required Course #2": 4, "Courses Satisfied": 4}
    required_course_two_data = ["ISE 321"]

    # Same row_limit
    nextSectionRowStarter2 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert2, required_course_two_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter2[1]


    # Same row_limit
    row_limit_two = 2
    valuesToPopAndInsert3 = {"Required Course #3 (ONE course)": 4, "Courses Satisfied": 4}
    required_course_data_three = ["ISE 331", "CSE 331"] # They are separate courses


    nextSectionRowStarter3 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter2[0], column_value_start, column_value_end, row_limit_two, valuesToPopAndInsert3, required_course_data_three, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter3[1]


    # Same row_limit
    valuesToPopAndInsert4 = {"Required Course #4 (ONE course)": 4, "Courses Satisfied": 4}
    required_course_data_four = ["CSE/ISE 337"] 

    nextSectionRowStarter4 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter3[0], column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert4, required_course_data_four, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter4[1]

    # Same row_limit

    valuesToPopAndInsert5 = {"Required Course #5 (ONE course)": 4, "Courses Satisfied": 4}
    required_course_data_five = ["ISE 475, 488"]
    
    nextSectionRowStarter5 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter4[0], column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert5, required_course_data_five, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter5[1]

    # New row_limit
    valuesToPopAndInsert6 = {"Required Course #6 (ONE course)": 4, "Courses Satisfied": 4}
    required_course_data_six = ["ESE 422", "CSE 370", "EST 393", "BUS 393, 346", "AMS 341"]

    row_limit_new = 5
    nextSectionRowStarter6 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter5[0], column_value_start, column_value_end, row_limit_new, valuesToPopAndInsert6, required_course_data_six, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter6[1]

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter6[0], column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus  

# Outlines the tech and systems table and displays all the required classes needed to take   
@staticmethod
def techAndSystemsTable(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:

    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Technological Systems Management Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    valuesToPopAndInsert1 = {"Required Course #1 (ONE course)": 4, "Courses Satisfied": 4}
    required_course_one_data = ["EST 201, 202"]

    required_starter = row_value_start + 1
    row_limit_one = 1
    
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, required_starter, column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert1, required_course_one_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    valuesToPopAndInsert2 = {"Required Course #2": 4, "Courses Satisfied": 4}
    required_course_two_data = ["EST 391"]

    # Same row_limit
    nextSectionRowStarter2 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert2, required_course_two_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter2[1]


    # Same row_limit
    valuesToPopAndInsert3 = {"Required Course #3": 4, "Courses Satisfied": 4}
    required_course_data_three = ["EST 392"]


    nextSectionRowStarter3 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter2[0], column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert3, required_course_data_three, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter3[1]


    # Same row_limit
    valuesToPopAndInsert4 = {"Required Course #4": 4, "Courses Satisfied": 4}
    required_course_data_four = ["EST 393"]

    nextSectionRowStarter4 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter3[0], column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert4, required_course_data_four, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter4[1]

  
    valuesToPopAndInsert5 = {"2 of the following courses": 4, "Courses Satisfied": 4}
    required_course_data_five = ["ISE 323", "ISE 340/EST 310", "EST 320, 323, 326, 327, 364"]
    
    row_limit_new = len(required_course_data_five) 
    nextSectionRowStarter5 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter4[0], column_value_start, column_value_end, row_limit_new, valuesToPopAndInsert5, required_course_data_five, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter5[1]

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter5[0], column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus  

# Outlines the bfinance systems table and displays all the required classes needed to take   
@staticmethod
def financeSystemsTable(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:

    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Financial Information Systems Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    valuesToPopAndInsert1 = {"2 of the following courses": 4, "Courses Satisfied": 4}
    required_course_one_data = ["CSE 215", "AMS 315, 318"]

    required_starter = row_value_start + 1
    row_limit_one = 2
    
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, required_starter, column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert1, required_course_one_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    valuesToPopAndInsert2 = {"4 of the following courses": 4, "Courses Satisfied": 4}
    required_course_two_data = ["ACC 210", "AMS 311, 316, 320, 341, 394, 441", "BUS 330, 331, 355, 356", "CSE/ISE/EST 323", "ISE 331"]

    # Same row_limit
    row_limit_new = 5
    nextSectionRowStarter2 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, row_limit_new, valuesToPopAndInsert2, required_course_two_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter2[1]

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter2[0], column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus 

# Outlines the health informatics table and displays all the required classes needed to take   
@staticmethod
def healthInformaticsTable(sheet, row_value_start, column_value_start, column_value_end, class_amount_needed, specializeCourses) -> str:
    
    total_courses_required_counter = 0

    sheet.merge_cells(start_row=row_value_start, start_column=column_value_start, end_row=row_value_start, end_column=column_value_end)
    sheet.cell(row=row_value_start, column=column_value_start).value = "Health Informatics Specialization"

    column_counter = column_value_start

    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    valuesToPopAndInsert1 = {"Required Course #1": 4, "Courses Satisfied": 4}
    required_course_one_data = ["HAN 200"]

    required_starter = row_value_start + 1
    row_limit_one = 1
    
    nextSectionRowStarter1 = specializationTableInputterAlgorithm(sheet, required_starter, column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert1, required_course_one_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter1[1]

    # Then we might wanna do a check as well to see if the course is actually there or not 
    valuesToPopAndInsert2 = {"Required Course #2 (ONE course)": 4, "Courses Satisfied": 4}
    required_course_two_data = ["BIO 202, 203"]

    # Same row_limit
    nextSectionRowStarter2 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter1[0], column_value_start, column_value_end, row_limit_one, valuesToPopAndInsert2, required_course_two_data, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter2[1]

    valuesToPopAndInsert3 = {"4 of the following courses": 4, "Courses Satisfied": 4}
    required_course_data_three = ["PSY 103", "BME 205", "HAN 202", "CSE 377", "ECO 327", "BCP 405"]

    row_limit_new = 6
    nextSectionRowStarter3 = specializationTableInputterAlgorithm(sheet, nextSectionRowStarter2[0], column_value_start, column_value_end, row_limit_new, valuesToPopAndInsert3, required_course_data_three, specializeCourses)
    total_courses_required_counter += nextSectionRowStarter3[1]

    last_row_limit = -1
    completedStatus = specalizationCompleteQuestionLine(sheet, nextSectionRowStarter3[0], column_value_start, column_value_end, last_row_limit, total_courses_required_counter, class_amount_needed)
    return completedStatus 

# This function is responsible for creating the table as well as inputting the required classes that are needed to satisfy
# a particular specialization. In addition, if a course that a student takes satisfies one of the required courses, then
# that course will be noted down next to the required course that was satisfied.
@staticmethod
def specializationTableInputterAlgorithm(sheet, row_start, column_start, column_end, row_limit, insert_dictionary, required_course_data, specializeCourses) -> list:

    class_sum_to_return = 0 

    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, '': 0.00}

    sheet.merge_cells(start_row=row_start, start_column=column_start, end_row=row_start, end_column=column_start+3) # For The First 3 columns
    sheet.merge_cells(start_row=row_start, start_column=column_start+4, end_row=row_start, end_column=column_end) # For the last 2 columns

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    initial_column = column_start 
    keys_in_dict = list(insert_dictionary.keys())
    while len(insert_dictionary) != 0:
        sheet.cell(row=row_start, column=initial_column).value = keys_in_dict[0]
        sheet.cell(row=row_start, column=initial_column).font = Font(size=10, name='Calibri', bold=True)
        valueOffset = insert_dictionary[keys_in_dict[0]]
        del insert_dictionary[keys_in_dict[0]]
        keys_in_dict.pop(0)
        initial_column += valueOffset
    
    initial_column = column_start
    while initial_column != column_end + 1:
        sheet.cell(row=row_start, column=initial_column).border = all_sides_border
        initial_column += 1
    
    row_lower = row_start + 1
    row_max = row_lower + row_limit

    # Responsible for injecting the preassumed data 
    keys_of_courses = list(required_course_data)

    while row_lower != row_max: # Injects some presumed data 
        sheet.merge_cells(start_row=row_lower, start_column=column_start, end_row=row_lower, end_column=column_start+3)
        sheet.merge_cells(start_row=row_lower, start_column=column_start+4, end_row=row_lower, end_column=column_end)
        try:
            keyFirst = keys_of_courses[0]
            courseName = keyFirst
            sheet.cell(row=row_lower, column=column_start).value = courseName
        except: pass
        sheet.cell(row=row_lower, column=column_start).font = Font(size=10)
        sheet.cell(row=row_lower, column=column_start).border = _border_style_left # Border to Lower Division
        sheet.cell(row=row_lower, column=column_start+3).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+4).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+5).border = _border_style_left
        sheet.cell(row=row_lower, column=column_start+6).border = _border_style_left
        sheet.cell(row=row_lower, column=column_end-1).border = _border_style_left
        sheet.cell(row=row_lower, column=column_end).border = _border_style_right
        try:
            keys_of_courses.pop(0)
        except:
            pass
        row_lower += 1
    
    # Before returning, we now check to see if the course exists in our specializeCourses dictionary 
    getListOfSortedKeys = list(sorted(specializeCourses.keys()))
    lengthOfList = len(getListOfSortedKeys)

    searching_start_row = row_start + 1
    row_offset = 1

    while lengthOfList != 0:

        getFirstIndex = getListOfSortedKeys[0]

        if searching_start_row >= row_max: # Not found, so we pop this key from the list and continue 

            searching_start_row = row_start + 1
            row_offset = 1
            getListOfSortedKeys.pop(0)
            lengthOfList -= 1
            continue # So we can start looking at the next element in the list right away 
        
        
        cellValue = sheet.cell(row=searching_start_row, column=column_start).value

        # Need a different way to evaluate 
        boolISECSE, boolISEEST, boolCSEISEEST = False, False, False  
        if "CSE/ISE" in getFirstIndex:
            if getFirstIndex[0:3] in cellValue and getFirstIndex[4:7] in cellValue:
                boolISECSE = True
        
        elif "ISE 340/EST 310" in getFirstIndex:
            if getFirstIndex[0:3] in cellValue and getFirstIndex[8:11] in cellValue:
                boolISEEST = True 
        
        elif "CSE/ISE/EST 323" in getFirstIndex:
            if getFirstIndex[0:3] in cellValue and getFirstIndex[4:7] in cellValue and getFirstIndex[8:11]:
                boolCSEISEEST = True
            
        
        # First check the description if it is in the cellValue 
        if getFirstIndex[0:3] in cellValue or boolISECSE == True or boolISEEST == True or boolCSEISEEST == True: # Found the class, now check if the grade is of acceptable value 
            # Then check to see if the classNumber is also in the cellValue

            acceptDiff = False

            index1 = 4 if boolISECSE == False else 8 
            index2 = 7 if boolISECSE == False else 11

            if boolISEEST == True and getFirstIndex[4:7] in cellValue and getFirstIndex[12:15] in cellValue: acceptDiff = True

            index1 = index1 if boolCSEISEEST == False else 12
            index2 = index2 if boolCSEISEEST == False else 15

            if getFirstIndex[index1:index2] in cellValue or acceptDiff == True:

                getGrade = specializeCourses[getFirstIndex].grade # Get the grade of the object 
                if getGrade in acceptable_grades:
                    valueCorresponding = acceptable_grades[getGrade]
                else:
                    valueCorresponding = 0.00
                if float(valueCorresponding) == 0.00: # Then we don't accept the class so we instead go back
                    searching_start_row = row_start + 1 # Reset back to original spot and use a new class 
                    row_offset = 1
                    getListOfSortedKeys.pop(0)
                    lengthOfList -= 1      
                else: # Acceptable class
                    class_sum_to_return += 1
                    currentCellValue = sheet.cell(row=searching_start_row, column=column_start+4).value
                    if currentCellValue == None:
                        sheet.cell(row=searching_start_row, column=column_start+4).value = f'{getFirstIndex}'
                    else:
                        newCellValue = ""
                        if currentCellValue[0:3] == getFirstIndex[0:3]: # Same class Description
                            newCellValue = currentCellValue + ", " + f'{getFirstIndex[4:7]}' # We just put the class number
                        else: # Different class description
                            newCellValue = currentCellValue + ", " + f'{getFirstIndex}'
                        sheet.cell(row=searching_start_row, column=column_start+4).value = str(newCellValue)

                    searching_start_row = row_start + 1 # Reset back to original spot 
                    row_offset = 1
                    getListOfSortedKeys.pop(0)
                    lengthOfList -= 1           
            else:
                searching_start_row += 1
                row_offset += 1              
        else:
            searching_start_row += 1
            row_offset += 1


    return [row_max, class_sum_to_return] # Return a list


# This function displays a message as to whether or not the corresponding specialization was satisfied or not.
# The message will display specialization completed if the student has completed all the required courses. Otherwise, it will display not completed
@staticmethod
def specalizationCompleteQuestionLine(sheet, row_value, column_value_start, column_value_end, row_limit, total_class_count, class_amount_needed) -> str:

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    row_lower = row_value + 1
    row_max = row_lower + row_limit

    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_max, column=start_col_merge).border = all_sides_border
        start_col_merge += 1

    sheet.merge_cells(start_row=row_max, start_column=column_value_start, end_row=row_max, end_column=column_value_end)
    sheet.cell(row=row_max, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)

    if total_class_count >= class_amount_needed: # Means specialization completed
        sheet.cell(row=row_max, column=column_value_start).value = f'Status: Specialization Completed ({total_class_count} of {class_amount_needed})'
        return "Yes"
    else:
        sheet.cell(row=row_max, column=column_value_start).value = f'Status: Specialization Incomplete ({total_class_count} of {class_amount_needed})'
        return "No"





    

