"""
After all the information has been stored, the purpose of this module
is to create the XLSX document. This module creates two sheets:

One sheet contains at least 3 tables that outlines the basic student information,
the required major courses, the courses taken per semester by the student,
and (if any) specialization that the student is currently taking

The second sheet contains multiple tables that list the possible specializations
that are available for CSE and ISE students and also lists the classes
that are needed for that particular specialization

"""

# This needs to get seriously rewritten, which we will do after the semester 

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from tablecreator.CreateISESpecDocument import createISESpecializationDocument
from tablecreator.CreateCSESpecDocument import createCSESpecializationDocument
from tablecreator.CreateSingularSpecDocument import createSingularSpecDocument

# Global variables that are module-bounded and are primarily used for formatting the tables for visual appeal 
_border_style_bottom = Border(bottom=Side(style='thin'))
_border_style_left = Border(left=Side(style='thin'))
_border_style_right = Border(right=Side(style='thin'))

# This is the entry point into creating the document. This function creates two sheets and calls 3 functions, each of which is 
# responsible for making a table that is to be displayed in the output.
# This function returns the name of the tabular document that is created 
def createDocument(fileNameInput, studentInformation, upperDivisionCourses, lowerDivisionCourses, technicalCSECourses, mathRequiredCourses, scienceCourses, sbcCourses, classesPerSemester, specalizeISECourses, specializeCSECourses) -> list:
 
    workbook = openpyxl.Workbook()

    getMajor = studentInformation['Major']

    sheet = workbook.create_sheet(f'{getMajor} Req')
    sheet2 = workbook.create_sheet(f'{getMajor} Spec') # This Spec is optional for CSE majors but required for ISE majors 

    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    starterColumnForPlanningGrid = createStudentTableInformation(sheet, studentInformation)
        
    # We want this method to reflect both CSE and ISE majors at the same time without creating much of a hassle 
    starterColumnForSpecTable = createMajorRequirements(sheet, sheet2, starterColumnForPlanningGrid, studentInformation, upperDivisionCourses, lowerDivisionCourses, technicalCSECourses, mathRequiredCourses, scienceCourses, sbcCourses, specalizeISECourses, specializeCSECourses)

    if studentInformation['Spec'] != "": # Specialization Declared
        starterColumnForPlanning = createISEDeclaredTable(sheet, starterColumnForSpecTable, studentInformation, specalizeISECourses, specializeCSECourses, technicalCSECourses) # Where specialization information is stored in these 3 main dictionaries
        createPlanningGrid(sheet, starterColumnForPlanning, classesPerSemester)
    else: # Specialization Not Declared
        createPlanningGrid(sheet, starterColumnForSpecTable, classesPerSemester)


    fileName = fileNameInput + '.xlsx'
    return [workbook, fileName]

# This function is responsible for creating the basic student information table 
# Similar to the other functions, this function does both creation of the table
# and the inputting of information into the table once the table has been made
# 
# The function returns the column value that is to be used for the next table  
@staticmethod
def createStudentTableInformation(sheet, studentInformation) -> int:

    row_value_start = 1 # Value subject to change, depending on user 
    start_col = 1 # Subject to change
    end_col = 2 # Subject to change, depends on the user 

    column_counter = start_col # Subject to change 

    sheet.append(['Student Information'])
    sheet.merge_cells(start_row=row_value_start, start_column=start_col, end_row=row_value_start, end_column=end_col)
    
    while column_counter != end_col + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=start_col).font = Font(size=14, bold=True, name='Calibri')

    # Add data to the Student Information table
    # This needs to be manually changed if it is required to add more fields
    data = []
    if studentInformation['Major'] == "CSE":
        data = [
        ['Last Updated'],
        ['Name'],
        ['ID Number'],
        ['Requirement Term'],
        ['CSHP'],
        ['Cumulative GPA'],
        ['Cumulative Credits'],
        ['Upper Division Credits']
    ]
    elif studentInformation['Major'] == "ISE":
        data = [
        ['Last Updated'],
        ['Name'],
        ['ID Number'],
        ['Requirement Term'],
        ['Cumulative GPA'],
        ['Cumulative Credits'],
        ['Upper Division Credits']
    ]

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20

    # Add data to the table
    rowTrackerForData = 2 # We start adding from the second row and go downwards, while the end_col is always going to constant
    for row in data:
        sheet.append(row)
        getKey = row.pop(0)
        getStudentInfo = studentInformation[getKey]
        try:
            sheet.cell(row=rowTrackerForData, column=end_col).value = float(getStudentInfo)
            sheet.cell(row=rowTrackerForData, column=end_col).alignment = Alignment(horizontal='left')
        except:
            sheet.cell(row=rowTrackerForData, column=end_col).value = str(getStudentInfo)
            sheet.cell(row=rowTrackerForData, column=end_col).alignment = Alignment(horizontal='left')
        rowTrackerForData += 1

    # Apply formatting to the Student Information table
    formattingEndValue = 0
    if studentInformation['Major'] == "CSE":
        formattingEndValue = 11
    elif studentInformation['Major'] == "ISE":
        formattingEndValue = 10
    end_row_value = 0
    for i in range(1, formattingEndValue):
        end_row_value += 1
        sheet.cell(row=i, column=start_col).font = Font(bold=True)
        sheet.cell(row=i, column=start_col).alignment = Alignment(horizontal='left')
    sheet.cell(row=row_value_start, column=start_col).font = Font(size=14, bold=True, name='Calibri')
    end_row_value -= 1 # Decrease by 1 since loop will run an extra time

    # Add Border To The Table 

    bottom_and_left_border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

    sheet.cell(row=row_value_start, column=start_col).border = _border_style_right
    sheet.cell(row=end_row_value, column=start_col).border = _border_style_bottom
    sheet.cell(row=end_row_value, column=end_col).border = bottom_and_left_border

    formattingBorderValue = 0
    if studentInformation['Major'] == "CSE":
        formattingBorderValue = 9
    elif studentInformation['Major'] == "ISE":
        formattingBorderValue = 8
    counter = 2
    for i in range(counter, formattingBorderValue):
        sheet.cell(row=i, column=end_col).border = _border_style_right

    return end_col+2 # We will start at a second col
        

# Function is responsible for keeping track of the classes the student took per semester 
@staticmethod
def createPlanningGrid(sheet: any, starter_column: int, classesPerSemester: dict) -> int:
  
    row_value_start = 1 # Subject to change 
    col_value_start = starter_column # Subject to change
    col_value_end = col_value_start + 7 # Subject to change (be sure to keep this at +7 if I ever wanna change it again!!!)

    # Get the last index of the list
    last_key = list(classesPerSemester.keys())[-1]

    acceptable_grades = ('A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'XFER')

    # ********************* THERE IS A POSSIBLE INFINITE LOOP HERE SO WE NEED TO INVESTIGATE ***********************
    
    lengthOfClasses = len(classesPerSemester)

    priority_term = {'Fall': 'Winter', 'Winter': 'Spring', 'Spring': 'Summer', 'Summer': 'Fall'}

    while (lengthOfClasses % 4) != 0: 
           
        keyAtLastIndex = list(classesPerSemester)[-1] # Grab the last semester 
        # Once we get the key at the last index, we then want to parse this information 
        whiteSpaceIndex = keyAtLastIndex.index(' ')
        semesterTerm, semesterYear = keyAtLastIndex[0:whiteSpaceIndex], keyAtLastIndex[whiteSpaceIndex+1:]

        getNewTerm = priority_term[semesterTerm]

        if getNewTerm == 'Winter': 
            semesterYear = int(semesterYear) + 1 # Add to the semester Year

        listOfClassesInFakeSemester = [] # Fake classes, so empty list 
        classesPerSemester[f'{semesterTerm} {semesterYear}'] = listOfClassesInFakeSemester # We add no class information
        lengthOfClasses = lengthOfClasses + 1

        if (lengthOfClasses % 4) == 0: 
            break 

    # ********************* THERE IS A POSSIBLE INFINITE LOOP HERE SO WE NEED TO INVESTIGATE ******************************


    column_counter = col_value_start
    sheet.cell(row=row_value_start, column=col_value_start).value = "Courses Taken & Credits Earned"
    sheet.merge_cells(start_row=row_value_start, start_column = col_value_start, end_row=row_value_start, end_column=col_value_end)

    while column_counter != col_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=col_value_start).font = Font(size=14, bold=True, name='Calibri')


    highestListValue = 0

    for keys in classesPerSemester:
        getList = classesPerSemester[keys]
        if len(getList) > highestListValue:
            highestListValue = len(getList)
    highestListValue += 1

    starting_row = row_value_start + 1 # This only gets incremented when we hit starting_column value of 11
    starting_column = col_value_start # We increment this until we hit the value of K (which is the value of 11 )

    keys_list = list(classesPerSemester.keys())
    lengthOfListKeys = len(keys_list)        
    index = 0

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    recentClass = False

    while lengthOfListKeys != 0:

        getFirstIndex = keys_list[index]

        if getFirstIndex == last_key: 
            recentClass = True

        index += 1
        indexOfWhiteSpace = getFirstIndex.index(' ')
        seasonInformation, seasonYear = getFirstIndex[0:indexOfWhiteSpace], getFirstIndex[indexOfWhiteSpace+1:]


        if starting_column < col_value_end: 

            sheet.cell(row=starting_row, column=starting_column).value = seasonInformation
            sheet.cell(row=starting_row, column=starting_column).border = all_sides_border
            try: sheet.cell(row=starting_row, column=starting_column+1).value = float(seasonYear) 
            except: sheet.cell(row=starting_row, column=starting_column+1).value = str(seasonYear) 
            sheet.cell(row=starting_row, column=starting_column+1).border = all_sides_border

            # After we added the information, we now add the classes at the bottom
            getList = classesPerSemester[f'{seasonInformation} {seasonYear}']
            
            tempStartingRow = starting_row
            tempStartingCol = starting_column
            classCounter = 0
            creditsCounter = 0

            while classCounter != highestListValue:

                try:
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).value = getList[0].courseName if (recentClass != True and getList[0].grade != 'I') else getList[0].courseName + "*"
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).border = _border_style_left
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).value =  float(getList[0].credits) if (getList[0].grade in acceptable_grades) else (float(0) if str(getList[0].grade) == "" else str(getList[0].grade))
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).border = _border_style_right
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).alignment = Alignment(horizontal='right')

                    try:
                        creditsCounter += float(getList[0].credits)
                    except: pass 

                    getList.pop(0)
                    tempStartingRow += 1
                    classCounter += 1
                except:
                    
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).border = _border_style_left
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).border = _border_style_right
                    tempStartingRow += 1
                    classCounter += 1

            # Reached here when we are done filling out the classes, so now we need to put the total credits
            # We use the highestListValue for this 

            sheet.cell(row=starting_row+highestListValue, column=starting_column).value = "Total: "
            sheet.cell(row=starting_row+highestListValue, column=starting_column).border = all_sides_border
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).value = creditsCounter
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).border = all_sides_border

            #Make bold this part
            sheet.cell(row=starting_row+highestListValue, column=starting_column).font = Font(bold=True)
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).font = Font(bold=True)
            
            starting_column += 2
            lengthOfListKeys -= 1

        else: # This is the case when the starting column is greater than 11, which means we need to change our starting row value

            starting_row = starting_row + highestListValue + 1 # The next available space     
            starting_column = col_value_start # Put the starting column back to 4 

            sheet.cell(row=starting_row, column=starting_column).value = seasonInformation
            sheet.cell(row=starting_row, column=starting_column).border = all_sides_border
            sheet.cell(row=starting_row, column=starting_column+1).value = float(seasonYear)
            sheet.cell(row=starting_row, column=starting_column+1).border = all_sides_border

            getList = classesPerSemester[seasonInformation + " " + seasonYear]
            tempStartingRow = starting_row
            tempStartingCol = starting_column
            classCounter = 0
            creditsCounter = 0

            while classCounter != highestListValue:
                
                try:
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).value = getList[0].courseName if (recentClass != True and getList[0].grade != 'I') else getList[0].courseName + "*"
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).border = _border_style_left
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).value = float(getList[0].credits) if (getList[0].grade in acceptable_grades) else (float(0) if str(getList[0].grade) == "" else str(getList[0].grade))
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).border = _border_style_right
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).alignment = Alignment(horizontal='right')
                    creditsCounter += float(getList[0].credits)
                    getList.pop(0)
                    tempStartingRow += 1
                    classCounter += 1
                except:
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).border = _border_style_left
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).border = _border_style_right
                    tempStartingRow += 1
                    classCounter += 1
          
            sheet.cell(row=starting_row+highestListValue, column=starting_column).value = "Total: "
            sheet.cell(row=starting_row+highestListValue, column=starting_column).border = all_sides_border
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).value = creditsCounter
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).border = all_sides_border

            sheet.cell(row=starting_row+highestListValue, column=starting_column).font = Font(bold=True)
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).font = Font(bold=True)

            starting_column += 2
            lengthOfListKeys -= 1
    
    sheet.merge_cells(start_row=starting_row+highestListValue+1, start_column = col_value_start, end_row=starting_row+highestListValue+1, end_column=col_value_end)
    sheet.cell(row=starting_row+highestListValue+1, column=col_value_start).value = "* Means Course Is In-Progress (includes Incomplete courses)"
    sheet.cell(row=starting_row+highestListValue+1, column=col_value_start).font = Font(bold=True)


    return col_value_end+2 # 2 columns apart space 


# Function is responsible for creating the Major Requirements of a specific student 
@staticmethod
def createMajorRequirements(sheet, sheet2, starter_column, studentInformation, upperDivisionCourses, lowerDivisionCourses, technicalCSECourses, mathRequiredCourses, scienceCourses, sbcCourses, specalizeISECourses, specializeCSECourses) -> int:

    row_value_start = 1 # Subject to change
    column_value_start = starter_column # Subject to change
    column_value_end = column_value_start + 8 # Subject to change 

    if (column_value_end - column_value_start + 1) % 3 != 0:
        column_value_start += 2 # Shift by 2 
        column_value_end += 3

    column_counter = column_value_start

    if studentInformation['Major'] == "CSE":
        sheet.cell(row=row_value_start, column=column_value_start).value = "CSE BS Degree Requirements"
    elif studentInformation['Major'] == "ISE":
        sheet.cell(row=row_value_start, column=column_value_start).value = "ISE BS Degree Requirements"
    sheet.merge_cells(start_row=row_value_start, start_column = column_value_start, end_row=row_value_start, end_column=column_value_end)
    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    # We wanna have a bottom-down approach in making the table and let other functions be responsible for making our table

    upperDivisionCSERowStart = handleLowerDivisonCourses(sheet, row_value_start, column_value_start, column_value_end, studentInformation, lowerDivisionCourses) # Method Complete!
    technicalCSERowStart = handleUpperDivisionCourses(sheet, upperDivisionCSERowStart, column_value_start, column_value_end, studentInformation, upperDivisionCourses) # Method Complete!
    mathCoursesRowStart = handleTechnicalCourses(sheet, technicalCSERowStart, column_value_start, column_value_end, studentInformation, technicalCSECourses) # Method Works!
    scienceOrISECoursesRowStart = handleMathRequiredCourses(sheet, mathCoursesRowStart, column_value_start, column_value_end, studentInformation, mathRequiredCourses) # Method Works!

    sbcCoursesRowStartOrSpecRowStart = 0

    if studentInformation['Major'] == "CSE":
        sbcCoursesRowStartOrSpecRowStart = handleScienceRequiredCourses(sheet, scienceOrISECoursesRowStart, column_value_start, column_value_end, scienceCourses, studentInformation) # Don't necessarily need to pass in the student info but we will see
    elif studentInformation['Major'] == "ISE":
        sbcCoursesRowStartOrSpecRowStart = handleUpperDivisionWritingISE(sheet, scienceOrISECoursesRowStart, column_value_start, column_value_end, upperDivisionCourses, studentInformation)

    # We should probably combine the two dictionaries of technical and specialize courses
    if studentInformation['Major'] == "CSE":
        specializeCSECourses.update(technicalCSECourses) # Combine the two dictionaries 
    lineToHandleSbcsAfterwards = handleSpecalizationBoolean(sheet, sheet2, sbcCoursesRowStartOrSpecRowStart, column_value_start, column_value_end, specalizeISECourses, specializeCSECourses, studentInformation)
    handleSBCsCourses(sheet, lineToHandleSbcsAfterwards, column_value_start, column_value_end, studentInformation, sbcCourses)

    return column_value_end+2


# This function is responsible for creating the ISE specialization table that the student declared
@staticmethod
def createISEDeclaredTable(sheet, starter_column, studentInformation, specalizeISECourses, specializeCSECourses, technicalCSECourses):

    row_value_start = 1
    column_value_start = starter_column
    column_value_end = column_value_start + 8

    createSingularSpecDocument(sheet, row_value_start, column_value_start, column_value_end, studentInformation, specalizeISECourses, specializeCSECourses, technicalCSECourses)
    return column_value_end+2
        
# This function is responsible for performing sanitization on the lower division courses and calling a larger function
# that will create part of the table and input the lower division information afterwards 
@staticmethod
def handleLowerDivisonCourses(sheet, row_value_start, column_value_start, column_value_end, studentInformation, lowerDivisionCourses):

    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, '': 0.00}

    valuesToPopAndInsert = {'Lower Division Courses': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}

    # Now this is where we differ in our program and we may also have to consider honors program but for now
    # we will focus on CSE/ISE non-honors
    row_start = 0
    row_limit = 0
    temp_list_data = []
    lower_division_data = []

    # There are some alternate options for CSE majors
    flag150, flag160, flag260 = False, False, False

    if studentInformation['Major'] == "CSE":

        alternate_courses = ["CSE 150", "CSE 160", "CSE 260"] # These classes are only offered to CSE Honors Students  

        # Look through the lower division courses to see if these classes are there and if they are, adjust

        for value in alternate_courses:
            if value in lowerDivisionCourses.keys(): # Check the grade of this course 
                getGrade = lowerDivisionCourses[value].grade
                try:
                    valueCorresponding = acceptable_grades[getGrade]
                except:
                    valueCorresponding = 0.00
                if float(valueCorresponding) != 0.00: # Acceptable class 
                    if value == "CSE 150": 
                        flag150 = True
                    if value == "CSE 160":
                        # Check for 161 class credit
                        get161Grade = lowerDivisionCourses['CSE 161'].grade
                        try:
                            valueCorresponding = acceptable_grades[get161Grade]
                        except:
                            valueCorresponding = 0.00
                        if float(valueCorresponding) != 0.00: flag160 = True
                    if value == "CSE 260": 
                        get261Grade = lowerDivisionCourses['CSE 261'].grade
                        try:
                            valueCorresponding = acceptable_grades[get261Grade]
                        except:
                            valueCorresponding = 0.00
                        if float(valueCorresponding) != 0.00: flag260 = True
                        
    string150 = "CSE 150" if flag150 == True else "CSE 215"
    string160 = "CSE 160" if flag160 == True else "CSE 114 (TECH)"
    string161 = "CSE 161" if flag160 == True else ""
    string260 = "CSE 260" if flag260 == True else "CSE 214"
    string216 = "" if flag260 == True else "CSE 216"
    string261 = "CSE 261" if flag260 == True else ""


    if studentInformation['Major'] == "CSE":
        row_start = row_value_start + 1
        temp_list_data = [f'{string160}', f'{string161}', f'{string260}', f'{string261}', f'{string150}', f'{string216}', "CSE 220"]
        # Now remove the list of empty strings
        lower_division_data = [x for x in temp_list_data if x != ""] # List comprehension
        row_limit = len(lower_division_data)
    elif studentInformation['Major'] == "ISE": # ISE students can't take the selected honors classes so leave this as it is 
        row_start = row_value_start + 1
        row_limit = 3 # Will always be like that
        lower_division_data = ['CSE 114 (TECH)', 'CSE 214', 'ISE 218']

    returnNextRowStarter = excelFileInputterAlgorithm(sheet, row_start, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, lower_division_data, lowerDivisionCourses)
    return returnNextRowStarter

# This function is responsible for performing sanitization on the upper division courses and calling a larger function
# that will create part of the table and input the upper division information afterwards 
@staticmethod
def handleUpperDivisionCourses(sheet, upper_start, column_value_start, column_value_end, studentInformation, upperDivisionCourses):

    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, '': 0.00}

    valuesToPopAndInsert = {'Upper Division Courses': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}
    getCSHPValue = "" 
    upper_division_data = []

    # Who knows if non-honors CSE students decide to take honors classes, as this is possible to happen 

    flag350, flag385 = False, False
    flag316, flag320 = False, False

    if studentInformation['Major'] == "CSE":

        alternate_courses = ["CSE 350", "CSE 385"]

        for value in alternate_courses:
            if value in upperDivisionCourses.keys(): # Check the grade of this course 
                getGrade = upperDivisionCourses[value].grade
                try:
                    valueCorresponding = acceptable_grades[getGrade]
                except:
                    valueCorresponding = 0.00
                if float(valueCorresponding) != 0.00: # Acceptable class 
                    if value == "CSE 350": 
                        flag350 = True
                    if value == "CSE 385":
                        flag385 = True
    
    elif studentInformation['Major'] == "ISE":

        alternate_courses = ["ISE 316", "ISE 320"]

        for value in alternate_courses:
            if value in upperDivisionCourses.keys():
                getGrade = upperDivisionCourses[value].grade
                try:
                    valueCorresponding = acceptable_grades[getGrade]
                except:
                    valueCorresponding = 0.00
                if float(valueCorresponding) != 0.00:
                    if value == "ISE 316":
                        flag316 = True
                    if value == "ISE 320":
                        flag320 = True

    string310 = "CSE 310" if flag316 == False else "ISE 316"
    string340 = "BUS 340 (TECH)" if flag320 == False else "ISE 320"

    row_limit = 0

    if studentInformation['Major'] == "CSE": # Need to check if the student is part of CSHP 
        getCSHPValue = studentInformation['CSHP']
        if getCSHPValue == "NO": 
            row_limit = 8
        else: 
            row_limit = 10

        if flag350 == True and flag385 == True:
            upper_division_data = ['CSE/ISE 300 (SPK/WRTD)', 'CSE 350', 'CSE 310', 'CSE/ISE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE 385', 'CSE 416 (ESI/EXP+/SBS+/STEM+)'] if getCSHPValue == "NO" else ['CSE/ISE 300 (SPK/WRTD)', 'CSE 350', 'CSE 310', 'CSE/ISE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE 385', 'CSE 416 (ESI/EXP+/SBS+/STEM+)', 'CSE 495', 'CSE 496']
        elif flag350 == False and flag385 == True:
            upper_division_data = ['CSE/ISE 300 (SPK/WRTD)', 'CSE 303', 'CSE 310', 'CSE/ISE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE 385', 'CSE 416 (ESI/EXP+/SBS+/STEM+)'] if getCSHPValue == "NO" else ['CSE/ISE 300 (SPK/WRTD)', 'CSE 303', 'CSE 310', 'CSE/ISE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE 385', 'CSE 416 (ESI/EXP+/SBS+/STEM+)', 'CSE 495', 'CSE 496']
        elif flag350 == True and flag385 == False:
            upper_division_data = ['CSE/ISE 300 (SPK/WRTD)', 'CSE 350', 'CSE 310', 'CSE/ISE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE/MAT 373', 'CSE 416 (ESI/EXP+/SBS+/STEM+)'] if getCSHPValue == "NO" else ['CSE/ISE 300 (SPK/WRTD)', 'CSE 350', 'CSE 310', 'CSE/ISE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE/MAT 373', 'CSE 416 (ESI/EXP+/SBS+/STEM+)', 'CSE 495', 'CSE 496']
        else:
            upper_division_data = ['CSE/ISE 300 (SPK/WRTD)', 'CSE 303', 'CSE 310', 'CSE/ISE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE/MAT 373', 'CSE 416 (ESI/EXP+/SBS+/STEM+)'] if getCSHPValue == "NO" else ['CSE/ISE 300 (SPK/WRTD)', 'CSE 303', 'CSE 310', 'CSE/ISE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE/MAT 373', 'CSE 416 (ESI/EXP+/SBS+/STEM+)', 'CSE 495', 'CSE 496']
    
    elif studentInformation['Major'] == "ISE": # No CSHP for ISE students
        row_limit = 4 
        upper_division_data = [f'CSE/ISE 312 (CER/ESI/STAS)', f'ISE 305 (EXP+/TECH)', f'{string310}', f'{string340}'] # Also look for the course equivalents while inputting the information 
    
    returnNextSectionRowStarter = excelFileInputterAlgorithm(sheet, upper_start, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, upper_division_data, upperDivisionCourses)
    return returnNextSectionRowStarter

# This function is responsible for  calling a larger function that will create part of the table and input technical courses afterwards 
@staticmethod
def handleTechnicalCourses(sheet, technical_row_starter, column_value_start, column_value_end, studentInformation, technicalCSECourses):

    valuesToPopAndInsert = {'Technical Courses': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}

    row_limit = 4 # Let's keep this at 4 to keep things consistent 

    nextSectionRowStarter = excelFileInputterAlgorithm(sheet, technical_row_starter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, [], technicalCSECourses)
    return nextSectionRowStarter


# This function is responsible for performing sanitization on the math require courses and calling a larger function
# that will create part of the table and input math required courses afterwards 
@staticmethod
def handleMathRequiredCourses(sheet, math_start, column_value_start, column_value_end, studentInformation, mathRequiredCourses):

    requiredCourses = {}

    if studentInformation['Major'] == "CSE":
        requiredCourses = {'Calc 1': ["AMS 151 (QPS)", "MAT 131 (QPS)"], 'Calc 2': ["AMS 161 (QPS)", "MAT 132 (QPS)"], 'Lin Algebra': ["MAT 211 (STEM+)", "AMS 210 (STEM+)"], 'Graph Theory': ["AMS 301 (STEM+)"], 'Stats': ["AMS 310 (STEM+)", "AMS 311 (STEM+)"]}
    elif studentInformation['Major'] == "ISE":
        requiredCourses = {'Calc 1': ["AMS 151 (QPS)", "MAT 131 (QPS)"], 'Calc 2': ["AMS 161 (QPS)", "MAT 132 (QPS)", "CSE 215"], 'Lin Algebra': ["MAT 211 (STEM+)", "AMS 210 (STEM+)"], 'Stats': ["AMS 310 (STEM+)", "AMS 110", "ECO 320"]}

    math_copy = mathRequiredCourses.copy()

    math_final = {}

    # Part of a 3-course sequence, so we want to handle them independently
    if 'MAT 125' in math_copy.keys():
        math_copy['MAT 125'].courseName = 'MAT 125 (QPS)'
        math_final['MAT 125 (QPS)'] = math_copy['MAT 125']
    if 'MAT 126' in math_copy.keys():
        math_copy['MAT 126'].courseName = 'MAT 126 (QPS)'
        math_final['MAT 126 (QPS)'] = math_copy['MAT 126']
    if 'MAT 127' in math_copy.keys():
        math_copy['MAT 127'].courseName = 'MAT 127 (QPS)'
        math_final['MAT 127 (QPS)'] = math_copy['MAT 127']

    for keys in requiredCourses.keys():
        value_max = 1
        one_class_only = 0
        getList = requiredCourses[keys]
        for values in math_copy.keys():
            for items in getList:
                if values in items and one_class_only < value_max: 
                    math_copy[values].courseName = items
                    math_final[items] = math_copy[values]
                    one_class_only += 1
        

    valuesToPopAndInsert = {'Math Required Courses': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}

    row_limit = 0
    if 'MAT 125' in mathRequiredCourses.keys(): # This implies that the student may also take '126' (we might change back to '127' but we will see)
        if studentInformation['Major'] == "CSE":
            row_limit = 6 
        elif studentInformation['Major'] == "ISE":
            row_limit = 5
    else:
        if studentInformation['Major'] == "CSE":
            row_limit = 5
        elif studentInformation['Major'] == "ISE":
            row_limit = 4


    nextSectionRowStarter = excelFileInputterAlgorithm(sheet, math_start, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, [], math_final)
    return nextSectionRowStarter

# This function is responsible for performing sanitization on the science required courses and calling a larger function
# that will create part of the table and input the science required courses afterwards 
@staticmethod
def handleScienceRequiredCourses(sheet, row_starter, column_value_start, column_value_end, scienceCourses, studentInformation):

    valuesToPopAndInsert = {'Science Courses (9 Credits, 1 Lab)': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}

    labCombos = {'BIO 201': 'BIO 204', 'BIO 202': 'BIO 204', 'BIO 203': 'BIO 204', 'CHE 131': 'CHE 133', 'CHE 152': 'CHE 154', 'PHY 126': 'PHY 133', 'PHY 131': 'PHY 133', 'PHY 141': 'PHY 133'}
    independentCourses = ('AST 203', 'AST 205', 'CHE 132', 'CHE 321', 'CHE 322', 'CHE 331', 'CHE 332', 'GEO 102', 'GEO 103', 'GEO 112', 'GEO 113', 'GEO 122', 'PHY 125', 'PHY 127', 'PHY 132', 'PHY 134', 'PHY 142', 'PHY 251', 'PHY 252')

    science_copy = scienceCourses.copy()
    labClassFound = []
    laboratoryComboFound = False 
    for keys in science_copy:
        if keys in labCombos:
            labClassFound.append(keys)
            labDescription = labCombos[keys] # The class that we need
            list_of_keys = list(scienceCourses.keys())
            if laboratoryComboFound == False:
                for originalKeys in list_of_keys:
                    if labDescription in originalKeys:
                        laboratoryComboFound = True 
                        break
            if laboratoryComboFound == True:
                break # We found a class/lab combo, so we break out of the loop and do nothing 
    
    if laboratoryComboFound == False: # We found the class, but couldn't find the associated lab combo
        for keys in scienceCourses:
            try: 
                retrieveClass = labClassFound[0] 
                labClass = labCombos[retrieveClass]
                if retrieveClass in keys:
                    scienceCourses[keys].courseName = f'{retrieveClass} ({labClass} REQUIRED.)'
                labClassFound.pop(0)
            except: pass

        
    # Clear out any extraneous classes we may have 

    for keys in science_copy:
        if keys not in independentCourses:
            if (keys in labCombos) or (keys in labCombos.values()): pass # We then check the other dictionary pass
            else: del scienceCourses[keys]
    
    # Now we have three cases for the row limit:
    # 1) After iterating through the number of credits, if the sum isn't 9, the default row_limit is 4
    # 2) If the number of credits exceeds or is greater than 9, the default row_limit is how many times the iteration went
    # 3) If the number of credits exceeds or is greater than 9, but no lab combo, then the row limit is going to be number of iterations + 1

    # Go through credits
    creditsCounter = 0
    numberOfIterations = 0
    for keys in scienceCourses:
        creditsCounter += float(scienceCourses[keys].credits)
        numberOfIterations += 1

    row_limit = 0
    if laboratoryComboFound == False and creditsCounter >= 9: row_limit = numberOfIterations + 1
    elif laboratoryComboFound == True and creditsCounter >= 9: row_limit = numberOfIterations
    elif creditsCounter < 9: row_limit = 4 # Even if the lab is found or not found, if the credits is less than 4, then the row_limit is set to 4

    nextSectionRowStarter = excelFileInputterAlgorithm(sheet, row_starter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, [], scienceCourses)
    return nextSectionRowStarter

# This function is responsible for creating a separate table for handling the ISE upper division writing requirement
@staticmethod
def handleUpperDivisionWritingISE(sheet, writing_start, column_value_start, column_value_end, upperDivisionCourses, studentInformation):
    
    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, '': 0.00}

    writing_division_data = []

    alternate_courses = ["EST 304"]
    flag304 = False 

    for value in alternate_courses:
        if value in upperDivisionCourses.keys():
            getGrade = upperDivisionCourses[value].grade
            try:
                valueCorresponding = acceptable_grades[getGrade]
            except:
                valueCorresponding = 0.00
            if float(valueCorresponding) != 0.00:
                flag304 = True 
    
    valuesToPopAndInsert = {"Upper Division Writing": 3, "Grade": 1, "Credits": 1, "Term": 1, "Year": 1, "Comments": 2}

    if flag304 == False:
        writing_division_data = ['CSE/ISE 300 (SPK/WRTD)']
    else:
        writing_division_data = ['EST 304']

    #Only need to parse 1 class, so row_limit is going to be set to 1
    row_limit = 1 # For only 1 class 
    nextSectionRowStarter = excelFileInputterAlgorithm(sheet, writing_start, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, writing_division_data, upperDivisionCourses) # Hopefully this will put the information
    return nextSectionRowStarter 


# This function is responsible for calling either a CSE Specialization function or an ISE Specialization function (depending on the student's major)
# After either one of those functions return, we either get the name of the specialization the student declared or we did not
#
# In the event that we did get the name of the specialization that the student declared, we then call another method which
# like the other table creating methods, is responsible for creating a table that displays the student's progress of the specialization
@staticmethod
def handleSpecalizationBoolean(sheet, sheet2, sbcCoursesRowStartOrSpecRowStart, column_value_start, column_value_end, specalizeISECourses, specializeCSECourses, studentInformation):
    
    # Now this is where we are gonna call the function in the CreateISESpecDocument to create the other sheet
    specializationName = ""

    if studentInformation['Major'] == "ISE":
        specializationName = createISESpecializationDocument(sheet2, specalizeISECourses)
        if specializationName == "" and studentInformation['Spec'] == "": # Student did not declare specialization
            specializationName = "Specialization Not Declared"
        elif specializationName == "" and studentInformation['Spec'] != "": # Student declared specialization, but is not finished with it
            if studentInformation['Spec'] == "Network":
                specializationName = "Systems and Network Administration Specialiaztion Declared"
            elif studentInformation['Spec'] == "Health":
                specializationName = "Health Informatics Specialization Declared"
            elif studentInformation['Spec'] == "Economics":
                specializationName = "Business and Economics Specialization Declared"
            elif studentInformation['Spec'] == "Technological":
                specializationName = "Technological Systems Management Specialization Declared"
            elif studentInformation['Spec'] == "Financial":
                specializationName = "Financial Information Specialization Declared"
    elif studentInformation['Major'] == "CSE":
        specializationName = createCSESpecializationDocument(sheet2, specializeCSECourses)
        if specializationName == "" and studentInformation['Spec'] == "": # Student did not declare specialization
            specializationName = "Optional Specialization Not Declared"
        elif specializationName == "" and studentInformation['Spec'] != "": # Student declared specialization, but is not finished with it
            if studentInformation['Spec'] == "Artificial":
                specializationName = "Artificial Intelligence and Data Science Specialiaztion Declared"
            elif studentInformation['Spec'] == "Interaction":
                specializationName = "Human-Computer Interaction Specialization Declared"
            elif studentInformation['Spec'] == "Game":
                specializationName = "Game Programming Specialization Declared"
            elif studentInformation['Spec'] == "Security":
                specializationName = "Security and Privacy Specialization Declared"
            elif studentInformation['Spec'] == "System":
                specializationName = "Systems Software Development Specialization Declared"

    row_limit = -1
    nextSectionRowStarter = specializationISEInputterAlgorithm(sheet, sbcCoursesRowStartOrSpecRowStart, column_value_start, column_value_end, row_limit, specializationName)
    return nextSectionRowStarter

    # We are only going to include one line in the data table 


# This function is responsible  calling a larger function that will create part of the table and input the SBC courses afterwards 
@staticmethod
def handleSBCsCourses(sheet, row_starter, column_value_start, column_value_end, studentInformation, sbcCourses):

    valuesToPopAndInsert = {'SBC Objectives': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}
    sbc_data = []
    row_limit = 0

    if studentInformation['Major'] == "CSE":
        sbc_data = ['ARTS', 'GLO', 'HUM', 'SBS', 'USA', 'WRT', 'DIV']
        row_limit = 7
    elif studentInformation['Major'] == "ISE":
        sbc_data = ['ARTS', 'GLO', 'HUM', 'SBS', 'SNW', 'USA', 'WRT', 'SBS+', 'DIV']
        row_limit = 9
    
    sbcInputterAlgorithm(sheet, row_starter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, sbc_data, sbcCourses)

# This is the larger function that is responsible for creating parts of a table and inputting information into those parts 
@staticmethod
def excelFileInputterAlgorithm(sheet, row_value, column_value_start, column_value_end, row_limit, insert_dictionary: dict, class_data: dict, category_dictionary: dict):

    # These courses gets factored into GPA 
    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, 'F': 0.00, 'IF': 0.00, 'Q': 0.00}

    sheet.merge_cells(start_row=row_value, start_column=column_value_start, end_row=row_value, end_column=column_value_start+2) # For The First 3 columns
    sheet.merge_cells(start_row=row_value, start_column=column_value_end-1, end_row=row_value, end_column=column_value_end) # For The Last 2 columns

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    initial_column = column_value_start 
    keys_in_dict = list(insert_dictionary.keys())

    while len(insert_dictionary) != 0:
        sheet.cell(row=row_value, column=initial_column).value = keys_in_dict[0]
        sheet.cell(row=row_value, column=initial_column).font = Font(size=10, name='Calibri', bold=True)
        valueOffset = insert_dictionary[keys_in_dict[0]]
        del insert_dictionary[keys_in_dict[0]]
        keys_in_dict.pop(0)
        initial_column += valueOffset
    
    initial_column = column_value_start
    while initial_column != column_value_end + 1:
        sheet.cell(row=row_value, column=initial_column).border = all_sides_border
        initial_column += 1
    
    row_lower = row_value + 1
    row_max = row_lower + row_limit

    # Define our keys_of_courses depending on our class_data variable 
    keys_of_courses = list(category_dictionary.keys()) if not class_data else list(class_data)
  

    # Injects the class information into the cells: just the class CourseName though 
    while row_lower != row_max:
         # Injects some presumed data 
        sheet.merge_cells(start_row=row_lower, start_column=column_value_start, end_row=row_lower, end_column=column_value_start+2)
        sheet.merge_cells(start_row=row_lower, start_column=column_value_end-1, end_row=row_lower, end_column=column_value_end)

        try:
            courseName = category_dictionary[keys_of_courses[0]].courseName if not class_data else keys_of_courses[0]
            sheet.cell(row=row_lower, column=column_value_start).value = courseName

        except: pass

        sheet.cell(row=row_lower, column=column_value_start).font = Font(size=10)
        sheet.cell(row=row_lower, column=column_value_start).border = _border_style_left # Border to Lower Division
        sheet.cell(row=row_lower, column=column_value_start+3).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+4).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+5).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+6).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_end-1).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_end).border = _border_style_right

        try:
            keys_of_courses.pop(0)
        except:
            pass
        row_lower += 1
    
    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_lower, column=start_col_merge).border = all_sides_border
        start_col_merge += 1


    # Begin inserting data. This is where we actually insert the class information
    getListOfSortedKeys = list(sorted(category_dictionary.keys()))
    lengthOfList = len(getListOfSortedKeys)
    
    creditsCounter, creditsValue = 0, 0

    searching_start_row = row_value + 1
    row_offset = 1

    # But before we actually begin, we need to do some computing
    while lengthOfList != 0:

        getFirstIndex = getListOfSortedKeys[0]

        if searching_start_row >= row_max: # Not found, so we pop this key from the list and continue 

            searching_start_row = row_value + 1
            row_offset = 1
            getListOfSortedKeys.pop(0)
            lengthOfList -= 1
            continue # So we can start looking at the next element in the list right away 
          
        cellValue = sheet.cell(row=searching_start_row, column=column_value_start).value

        # So here, we should also consider the case about alternative classes

        if getFirstIndex in cellValue: # Found the element  
                
                insertion_start = column_value_start + 3

                # Then we need to check and see if we have to replace the cell value 
                
                while insertion_start != column_value_end:

                    getColumnDescription = str(sheet.cell(row=searching_start_row-row_offset, column=insertion_start).value)
                    getColumnDescription = getColumnDescription.lower()
                    classObject = category_dictionary[getFirstIndex]
                    field_value = getattr(classObject, getColumnDescription)
                    try: 
                        sheet.cell(row=searching_start_row, column=insertion_start).value = float(field_value) # Put the grade of the object
                        if getColumnDescription == "credits":
                            # We should not include courses that has P, S, XFER, W
                            getGradeLetter = str(sheet.cell(row=searching_start_row, column=insertion_start-1).value)
                            if getGradeLetter == 'IF':
                                getGradeLetter = 'F'
                            # These courses are NOT factored into the GPA. This also includes courses that have an empty grade 
                            if getGradeLetter != 'XFER' and getGradeLetter != 'P' and getGradeLetter != 'S' and getGradeLetter != 'W' and getGradeLetter != 'U' and getGradeLetter != 'NC' and getGradeLetter != 'I' and getGradeLetter != '':
                                creditsCounter += float(field_value)
                                getPoints = acceptable_grades[getGradeLetter]
                                letterAndCreditProduct = float(getPoints) * float(field_value)
                                creditsValue += letterAndCreditProduct
                    except: 
                            sheet.cell(row=searching_start_row, column=insertion_start).value = str(field_value) # Put the grade of the object 
                    sheet.cell(row=searching_start_row, column=insertion_start).alignment = Alignment(horizontal='left')
                    insertion_start += 1

                searching_start_row = row_value + 1 # Reset back to original spot 
                row_offset = 1
                getListOfSortedKeys.pop(0)
                lengthOfList -= 1         
        else: # If we didnt find the element, we keep searching but using the same element
                searching_start_row += 1 # Increment by 1 
                row_offset += 1

    if creditsCounter != 0: totalGPA = round(float(creditsValue/creditsCounter), 3)
    else: totalGPA = round(float(0), 3)

    sheet.merge_cells(start_row=row_max, start_column=column_value_start, end_row=row_max, end_column=column_value_end)
    sheet.cell(row=row_max, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_max, column=column_value_start).value = f'Category GPA (excluding P, S, XFER, etc courses): {totalGPA}'
    return row_max+1 # Cause we need this information to be passed down and is also the next row that will hold the section information

# This function is for creating the ISE specialization table  
@staticmethod
def specializationISEInputterAlgorithm(sheet, row_value, column_value_start, column_value_end, row_limit, specializationName):
    
    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    row_lower = row_value + 1
    row_max = row_lower + row_limit

    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_max, column=start_col_merge).border = all_sides_border
        start_col_merge += 1

    sheet.merge_cells(start_row=row_max, start_column=column_value_start, end_row=row_max, end_column=column_value_end)
    sheet.cell(row=row_max, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_max, column=column_value_start).value = f'{specializationName}'
    return row_max+1

# This function is responsible for creating part of the SBC table that is also part of the major requirements 
@staticmethod
def sbcInputterAlgorithm(sheet, row_value, column_value_start, column_value_end, row_limit, insert_dictionary, class_data, category_dictionary):

    # List of acceptable grades that are factored into the GPA 
    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, 'F': 0.00, 'Q': 0.00}

    sheet.merge_cells(start_row=row_value, start_column=column_value_start, end_row=row_value, end_column=column_value_start+2) # For Lower Division Courses
    sheet.merge_cells(start_row=row_value, start_column=column_value_end-1, end_row=row_value, end_column=column_value_end) # For Comments

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    initial_column = column_value_start 
    keys_in_dict = list(insert_dictionary.keys())
    while len(insert_dictionary) != 0:
        sheet.cell(row=row_value, column=initial_column).value = keys_in_dict[0]
        sheet.cell(row=row_value, column=initial_column).font = Font(size=10, name='Calibri', bold=True)
        valueOffset = insert_dictionary[keys_in_dict[0]]
        del insert_dictionary[keys_in_dict[0]]
        keys_in_dict.pop(0)
        initial_column += valueOffset
    
    initial_column = column_value_start
    while initial_column != column_value_end + 1:
        sheet.cell(row=row_value, column=initial_column).border = all_sides_border
        initial_column += 1
    
    row_lower = row_value + 1
    row_max = row_lower + row_limit

    keys_of_courses = []
    if len(class_data) == 0: # Means no preassumed data
        keys_of_courses = list(category_dictionary.keys())
    else: # There is preassumed data 
        keys_of_courses = list(class_data)

    while row_lower != row_max:
        sheet.merge_cells(start_row=row_lower, start_column=column_value_start, end_row=row_lower, end_column=column_value_start+2)
        sheet.merge_cells(start_row=row_lower, start_column=column_value_end-1, end_row=row_lower, end_column=column_value_end)
        try:
            keyFirst = keys_of_courses[0]
            courseName = ""
            if len(class_data) == 0:
                courseName = category_dictionary[keyFirst].courseName # Injection
            else: # Means there is some pre assumed data
                courseName = keyFirst # Assumption 
            sheet.cell(row=row_lower, column=column_value_start).value = courseName
        except: pass
        sheet.cell(row=row_lower, column=column_value_start).font = Font(size=10)
        sheet.cell(row=row_lower, column=column_value_start).border = _border_style_left # Border to Lower Division
        sheet.cell(row=row_lower, column=column_value_start+3).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+4).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+5).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+6).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_end-1).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_end).border = _border_style_right
        try:
            keys_of_courses.pop(0)
        except:
            pass
        row_lower += 1
    
    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_lower, column=start_col_merge).border = all_sides_border
        start_col_merge += 1


    getListOfSortedKeys = list(category_dictionary.keys())
    lengthOfList = len(getListOfSortedKeys)
    creditsCounter = 0
    creditsValue = 0

    searching_start_row = row_value + 1
    row_offset = 1

    # This algorithm varies in a sense that the object contains a list of SBCs, so we need to insert them based on highest value 

    while lengthOfList != 0:    

        getFirstIndex = getListOfSortedKeys[0]
        getSBCClassList = category_dictionary[getFirstIndex].sbc # Will give us the sbcList

        if searching_start_row >= row_max: # Not found, so we pop this extraneous key from the list and continue 
            getListOfSortedKeys.pop(0) # Will pop the object even if there is an sbc that exists but is not in the required ones 
            searching_start_row = row_value + 1
            row_offset = 1
            lengthOfList -= 1

        cellValue = sheet.cell(row=searching_start_row, column=column_value_start).value

        if cellValue in getSBCClassList: # We found a course that has an sbc

            indexOfCellValue = category_dictionary[getFirstIndex].sbc.index(cellValue)
            sheet.cell(row=searching_start_row, column=column_value_start).value = f'{sheet.cell(row=searching_start_row, column=column_value_start).value} : {category_dictionary[getFirstIndex].courseName}'
            category_dictionary[getFirstIndex].sbc.pop(indexOfCellValue) 

            insertion_start = column_value_start + 3

            while insertion_start != column_value_end:
                getColumnDescription = str(sheet.cell(row=searching_start_row-row_offset, column=insertion_start).value)
                getColumnDescription = getColumnDescription.lower()
                classObject = category_dictionary[getFirstIndex]
                field_value = getattr(classObject, getColumnDescription)
                try: 
                    sheet.cell(row=searching_start_row, column=insertion_start).value = float(field_value) # Put the grade of the object
                    if getColumnDescription == "credits":
                        getGradeLetter = str(sheet.cell(row=searching_start_row, column=insertion_start-1).value)
                        if getGradeLetter == 'IF':
                            getGradeLetter = 'F' # We simply set it to 'F' if getGradeLetter is equal to IF, which means I/F
                        if getGradeLetter != 'XFER' and getGradeLetter != 'P' and getGradeLetter != 'S' and getGradeLetter != 'W' and getGradeLetter != 'U' and getGradeLetter != 'NC' and getGradeLetter != 'I' and getGradeLetter != '':
                            creditsCounter += float(field_value)
                            getPoints = acceptable_grades[getGradeLetter]
                            letterAndCreditProduct = float(getPoints) * float(field_value)
                            creditsValue += letterAndCreditProduct
                except:
                        sheet.cell(row=searching_start_row, column=insertion_start).value = str(field_value) # Put the grade of the object 
                sheet.cell(row=searching_start_row, column=insertion_start).alignment = Alignment(horizontal='left')
                insertion_start += 1

            if len(category_dictionary[getFirstIndex].sbc) == 0:
                getListOfSortedKeys.pop(0)
                lengthOfList -= 1
        else:
            searching_start_row += 1 # Increment by 1 
            row_offset += 1

    if creditsCounter != 0: totalGPA = round(float(creditsValue/creditsCounter), 3)
    else: totalGPA = round(float(0), 3)

    sheet.merge_cells(start_row=row_max, start_column=column_value_start, end_row=row_max, end_column=column_value_end)
    sheet.cell(row=row_max, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_max, column=column_value_start).value = f'Category GPA (excluding P, S, XFER, etc courses): {totalGPA}'
